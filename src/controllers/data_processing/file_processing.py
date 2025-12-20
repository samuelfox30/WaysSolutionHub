from openpyxl import load_workbook
from utils.logger import get_logger

logger = get_logger('file_processing')

def process_uploaded_file(file):
    logger.info("Iniciando processamento do arquivo de viabilidade")

    try:
        wb = load_workbook(file, data_only=True)
        ws = wb.active
        logger.info(f"Arquivo carregado com sucesso. Planilha ativa: {ws.title}")
    except Exception as e:
        logger.error(f"Erro ao carregar arquivo Excel: {str(e)}")
        raise Exception(f"Não foi possível carregar o arquivo Excel. Verifique se o arquivo está no formato correto. Erro: {str(e)}")

    def is_merged(cell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return True
        return False

    def is_blank_row(row_number):
        for cell in ws[row_number]:
            if cell.value not in (None, "", " "):
                return False
        return True

    def normalizar_percentual(valor):
        """
        Normaliza valores de porcentagem do Excel para o formato numérico correto.
        Excel armazena porcentagens como decimais (ex: 0.1327 = 13.27%)
        Multiplica por 100 para obter o valor percentual real.
        """
        if valor is None or valor == 0:
            return valor

        # Log do valor original para debug
        logger.debug(f"Normalizando percentual - Valor original: {valor} (tipo: {type(valor).__name__})")

        # Validar se o valor é numérico
        if isinstance(valor, str):
            logger.error(f"ERRO: Valor de percentual é TEXTO ao invés de número! Valor: '{valor}'")
            raise ValueError(f"A célula contém TEXTO ('{valor}') ao invés de um valor numérico. Verifique se a célula está formatada como número.")

        # Validar se é um tipo numérico válido
        if not isinstance(valor, (int, float)):
            logger.error(f"ERRO: Tipo de dado inválido para percentual! Tipo: {type(valor).__name__}, Valor: {valor}")
            raise ValueError(f"Tipo de dado inválido: {type(valor).__name__}. Esperado: número (int ou float).")

        # Excel armazena porcentagens como decimal, então multiplica por 100
        # Ex: 0.18313253012048192 * 100 = 18.313253012048192%
        valor_normalizado = valor * 100

        # Log do valor normalizado e verificação de limites
        if valor_normalizado > 99999999.99 or valor_normalizado < -99999999.99:
            logger.error(f"ERRO: Valor de percentual fora do limite do banco! Original: {valor}, Normalizado: {valor_normalizado}")
            logger.error(f"Limite do banco: DECIMAL(10,2) = -99999999.99 até 99999999.99")

        logger.debug(f"Valor normalizado: {valor_normalizado}")
        return valor_normalizado

    subgrupos = [
        "RECEITA",
        "CONTROLE DESPESAS POR NATURESAS SINTETICAS",
        "OBRIGAÇÕES",
        "GASTOS ADM",
        "MATERIA PRIMA",
        "GASTOS OPERACIONAIS",
        "PESSOAL"
    ]
    # Nota: "GASTOS OPERACIONAIS" aparece DUAS vezes: como bloco normal E como especial (veículos)

    # Colunas de cada cenário
    try:
        cenarios = [
            {"nome": ws["A1"].value, "cols": ("A", "B", "C")},
            {"nome": ws["E1"].value, "cols": ("E", "F", "G")},
            {"nome": ws["I1"].value, "cols": ("I", "J", "K")},
        ]
        logger.info(f"Cenários identificados: {[c['nome'] for c in cenarios]}")

        # Validar se os cenários foram identificados
        if not any(c["nome"] for c in cenarios):
            raise Exception("Não foi possível identificar os cenários (VIABILIDADE REAL, PE, IDEAL) na planilha. Verifique se o arquivo está no formato correto.")
    except Exception as e:
        logger.error(f"Erro ao identificar cenários: {str(e)}")
        raise

    # Detectar blocos nomeados (usando col A)
    blocos = {}
    for row in range(1, ws.max_row + 1):
        valor = ws[f"A{row}"].value
        if valor in subgrupos and is_merged(ws[f"A{row}"]):
            inicio = row + 1
            fim = inicio
            while fim <= ws.max_row and not is_blank_row(fim):
                fim += 1
            blocos[valor] = (inicio, fim - 1)

    logger.info(f"Blocos identificados: {list(blocos.keys())}")

    # Adicionar o bloco GERAL antes do primeiro subgrupo
    if blocos:
        primeiro_inicio = min(v[0] for v in blocos.values())
        blocos = {"GERAL": (3, primeiro_inicio - 2), **blocos}
    else:
        logger.warning("Nenhum bloco de subgrupo foi identificado na planilha")

    # Palavras-chave a ignorar
    ignorar_descricoes = {
        "RESULTADO REAL",
        "RESULTADO IDEAL",
        "PONTO DE EQUILIBRIO",
        "0"
    }

    # Extrair dados para cada cenário (os blocos normais)
    lista_cenarios = []
    total_itens_processados = 0

    for cenario in cenarios:
        nome_cenario = cenario["nome"]
        col_desc, col_perc, col_val = cenario["cols"]

        dados = {}
        itens_cenario = 0

        for nome, (ini, fim) in blocos.items():
            lista_itens = []
            for r in range(ini, fim + 1):
                desc = ws[f"{col_desc}{r}"].value
                perc = ws[f"{col_perc}{r}"].value
                valr = ws[f"{col_val}{r}"].value

                if desc is None:
                    continue

                desc_str = str(desc).strip().upper()
                if desc_str in ignorar_descricoes:
                    continue

                # Log detalhado do item sendo processado
                logger.debug(f"Linha {r} | Cenário: {nome_cenario} | Subgrupo: {nome} | Desc: {str(desc).strip()[:30]} | Perc original: {perc} | Valor: {valr}")

                # Normalizar o percentual
                try:
                    perc_normalizado = normalizar_percentual(perc)
                except ValueError as e:
                    # Erro de tipo de dado (string, etc)
                    logger.error(f"ERRO na linha {r}, coluna {col_perc}{r}")
                    logger.error(f"Cenário: {nome_cenario}, Subgrupo: {nome}, Descrição: {str(desc).strip()}")
                    logger.error(f"Valor do percentual: '{perc}' (tipo: {type(perc).__name__})")
                    raise Exception(f"Erro na linha {r}, coluna {col_perc}{r} (Descrição: '{str(desc).strip()}'): {str(e)}")
                except Exception as e:
                    logger.error(f"ERRO ao normalizar percentual na linha {r}, coluna {col_perc}{r}")
                    logger.error(f"Cenário: {nome_cenario}, Subgrupo: {nome}, Descrição: {str(desc).strip()}")
                    logger.error(f"Valor do percentual: {perc} (tipo: {type(perc).__name__})")
                    raise Exception(f"Erro inesperado na linha {r}, coluna {col_perc}{r}: {str(e)}")

                # Validar se o percentual normalizado está dentro dos limites
                if perc_normalizado is not None and (perc_normalizado > 99999999.99 or perc_normalizado < -99999999.99):
                    logger.error(f"ERRO: Percentual fora dos limites do banco de dados!")
                    logger.error(f"Linha {r}, Coluna {col_perc}{r} | Cenário: {nome_cenario} | Subgrupo: {nome}")
                    logger.error(f"Descrição: {str(desc).strip()}")
                    logger.error(f"Valor original: {perc}, Valor normalizado: {perc_normalizado}")
                    raise Exception(f"Valor de percentual fora dos limites na linha {r} (Descrição: '{str(desc).strip()}'): {perc_normalizado}. O valor deve estar entre -99999999.99 e 99999999.99")

                item = {
                    "cenario": nome_cenario,
                    "subgrupo": nome,
                    "descricao": str(desc).strip(),
                    "percentual": perc_normalizado,
                    "valor": valr
                }
                lista_itens.append(item)
                itens_cenario += 1
                total_itens_processados += 1
            dados[nome] = lista_itens

        logger.info(f"Cenário '{nome_cenario}': {itens_cenario} itens extraídos")
        lista_cenarios.append(dados)

    # ============================
    # Captura dos SUBGRUPOS ESPECIAIS
    # ============================

    especiais = {}

    # Procurar onde começa cada um
    especiais_nomes = [
        "DIVIDAS",
        "INVESTIMENTOS",
        "INVESTIMENTOS GERAL NO NEGOCIO",
        "GASTOS OPERACIONAIS"
    ]

    logger.info("Iniciando busca por subgrupos especiais...")
    print("\n" + "="*80)
    print("BUSCANDO SUBGRUPOS ESPECIAIS NA PLANILHA")
    print("="*80)

    for row in range(1, ws.max_row + 1):
        valor = ws[f"A{row}"].value
        if valor in especiais_nomes:
            inicio = row + 2  # pula uma linha após o título
            fim = inicio
            while fim <= ws.max_row and not is_blank_row(fim):
                fim += 1
            especiais[valor] = (inicio, fim - 1)
            logger.info(f"✓ Subgrupo especial '{valor}' encontrado na linha {row} (dados: linhas {inicio} a {fim - 1})")
            print(f"✓ Encontrado: '{valor}' na linha {row} | Dados: linhas {inicio} até {fim - 1}")

    # Log dos subgrupos especiais encontrados vs não encontrados
    print("\n" + "-"*80)
    print("RESUMO DOS SUBGRUPOS ESPECIAIS:")
    for nome_esperado in especiais_nomes:
        if nome_esperado in especiais:
            print(f"  ✓ {nome_esperado} - ENCONTRADO")
        else:
            print(f"  ✗ {nome_esperado} - NÃO ENCONTRADO")
            logger.warning(f"Subgrupo especial '{nome_esperado}' NÃO foi encontrado na planilha")
    print("-"*80 + "\n")

    # Extrair dados dos especiais
    dados_especiais = {}
    total_itens_especiais = 0

    print("\n" + "="*80)
    print("PROCESSANDO DADOS DOS SUBGRUPOS ESPECIAIS")
    print("="*80)

    for nome, (ini, fim) in especiais.items():
        print(f"\n>>> Processando: {nome} (linhas {ini} a {fim})")
        lista_itens = []
        for r in range(ini, fim + 1):
            desc = ws[f"A{r}"].value
            if desc is None:
                continue

            desc_str = str(desc).strip()
            if desc_str == "" or desc_str == "0":
                continue

            if nome in ("DIVIDAS", "INVESTIMENTOS"):
                parc = ws[f"B{r}"].value
                juros = ws[f"E{r}"].value
                total = ws[f"F{r}"].value

                # Ignorar se todos os valores forem 0 ou None
                if all(v in (None, 0, 0.0) for v in (parc, juros, total)):
                    continue

                item = {
                    "subgrupo": nome,
                    "descricao": desc_str,
                    "valor_parc": parc,
                    "valor_juros": juros,
                    "valor_total_parcela": total
                }

            elif nome == "INVESTIMENTOS GERAL NO NEGOCIO":
                val = ws[f"B{r}"].value
                if val in (None, 0, 0.0):
                    continue
                item = {
                    "subgrupo": nome,
                    "descricao": desc_str,
                    "valor": val
                }

            elif nome == "GASTOS OPERACIONAIS":
                custo_km = ws[f"B{r}"].value
                custo_mensal = ws[f"C{r}"].value

                print(f"  [GASTOS OP] Linha {r}: Desc='{desc_str}' | Custo KM={custo_km} | Custo Mensal={custo_mensal}")

                # Ignorar se ambos forem 0 ou None
                if all(v in (None, 0, 0.0) for v in (custo_km, custo_mensal)):
                    print(f"    └─> IGNORADO (ambos os valores são 0 ou None)")
                    continue

                item = {
                    "subgrupo": nome,
                    "descricao": desc_str,
                    "custo_km": custo_km,
                    "custo_mensal": custo_mensal
                }
                print(f"    └─> ADICIONADO ✓")

            lista_itens.append(item)
            total_itens_especiais += 1

        dados_especiais[nome] = lista_itens

        # Log especial para GASTOS OPERACIONAIS
        if nome == "GASTOS OPERACIONAIS":
            print(f"\n{'='*80}")
            print(f"RESULTADO GASTOS OPERACIONAIS: {len(lista_itens)} itens extraídos")
            if len(lista_itens) == 0:
                print("⚠️  ATENÇÃO: Nenhum item de GASTOS OPERACIONAIS foi extraído!")
                print("   Verifique se existem valores não-zero nas colunas B e C da seção GASTOS OPERACIONAIS")
            print(f"{'='*80}\n")

        logger.info(f"Subgrupo especial '{nome}': {len(lista_itens)} itens extraídos")

    # ============================
    # VALIDAÇÃO FINAL
    # ============================
    logger.info(f"Processamento concluído: {total_itens_processados} itens normais + {total_itens_especiais} itens especiais = {total_itens_processados + total_itens_especiais} itens totais")

    # Validar se algum dado foi processado
    if total_itens_processados == 0 and total_itens_especiais == 0:
        logger.error("FALHA: Nenhum dado foi extraído da planilha. Verifique se o formato está correto.")
        raise Exception("Nenhum dado foi extraído da planilha. Verifique se o arquivo está no formato esperado de Viabilidade Financeira.")

    if total_itens_processados == 0:
        logger.warning("AVISO: Nenhum item normal foi extraído, apenas itens especiais")

    # ============================
    # PRINTS PARA CONFERÊNCIA (DEBUG)
    # ============================
    logger.debug("Iniciando impressão dos dados extraídos para conferência")

    for dic in lista_cenarios:
        print("\n============================")
        print(f"CENÁRIO: {list(dic.values())[0][0]['cenario'] if list(dic.values())[0] else 'SEM NOME'}")
        for sub, itens in dic.items():
            print(f"\n--- {sub} ---")
            for i in itens:
                print(i)

    print("\n============================")
    print("SUBGRUPOS ESPECIAIS (únicos, replicar depois para os 3 cenários):")
    for sub, itens in dados_especiais.items():
        print(f"\n--- {sub} ---")
        for i in itens:
            print(i)

    logger.info("Arquivo processado com sucesso")
    return lista_cenarios, dados_especiais