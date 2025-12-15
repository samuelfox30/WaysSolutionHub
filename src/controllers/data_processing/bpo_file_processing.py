"""
BPO Financial Data Processing Module (Nova Estrutura)
=====================================================

Este módulo processa arquivos Excel com dados de BPO Financeiro (mensal).
Sheet: "Sheet"
Estrutura: Linha 1 = cabeçalho, Linha 2+ = dados

Autor: WaysSolutionHub
Data: 2025-11-25 (Refatorado)
"""

import openpyxl
from openpyxl import load_workbook
from utils.logger import get_logger

# Inicializar logger
logger = get_logger('bpo_processing')


# ============================================================================
# FUNÇÕES AUXILIARES
# ============================================================================

def extrair_codigo_e_nome(texto):
    """
    Extrai código hierárquico e nome de um texto como "1.01.06 - PMW RECEITA VENDA SERVIÇO"

    Returns:
        tuple: (codigo, nome, nivel_hierarquia)
        Exemplo: ("1.01.06", "PMW RECEITA VENDA SERVIÇO", 3)
    """
    texto = str(texto).strip()

    # Verificar se tem " - " separando código e nome
    if " - " in texto:
        partes = texto.split(" - ", 1)
        codigo = partes[0].strip()
        nome = partes[1].strip()
    else:
        # Se não tem separador, considera tudo como nome
        codigo = ""
        nome = texto

    # Calcular nível de hierarquia (contando pontos no código)
    nivel = codigo.count('.') + 1 if codigo else 0

    return codigo, nome, nivel


def converter_valor(valor):
    """Converte um valor de célula para float ou None"""
    if valor is None or valor == '':
        return None

    try:
        return float(valor)
    except (ValueError, TypeError):
        return None


def converter_porcentagem(valor):
    """Converte um valor de célula para porcentagem (multiplica por 100)"""
    if valor is None or valor == '':
        return None

    try:
        return float(valor) * 100
    except (ValueError, TypeError):
        return None


def extrair_meses_do_cabecalho(sheet):
    """
    Lê a linha 1 (cabeçalho) e extrai informações sobre os meses e colunas de totais.

    Formato esperado do cabeçalho:
    - Coluna A: Nome/Descrição
    - Colunas B+: "Mês Ano\nOrçado", "Mês Ano\nRealizado", etc (4 colunas por mês)
    - Últimas colunas: Contém a palavra "total" (ex: "Orçado total")

    Returns:
        dict: {
            'meses': [
                {'mes_nome': 'Janeiro', 'mes_numero': 1, 'ano': 2025, 'col_inicio': 2},
                ...
            ],
            'col_inicio_totais': 14,
            'num_meses': 3
        }
    """
    meses_info = []
    col_inicio_totais = None

    # Mapear nomes de meses para números
    meses_map = {
        'JANEIRO': 1, 'FEVEREIRO': 2, 'MARÇO': 3, 'MARCO': 3, 'ABRIL': 4,
        'MAIO': 5, 'JUNHO': 6, 'JULHO': 7, 'AGOSTO': 8,
        'SETEMBRO': 9, 'OUTUBRO': 10, 'NOVEMBRO': 11, 'DEZEMBRO': 12
    }

    total_colunas = sheet.max_column
    col_atual = 2  # Começa na coluna B (2)

    logger.debug("LENDO CABEÇALHO DA PLANILHA (Linha 1)")

    while col_atual <= total_colunas:
        cell_value = sheet.cell(row=1, column=col_atual).value

        if cell_value is None or str(cell_value).strip() == '':
            col_atual += 1
            continue

        cell_text = str(cell_value).strip().upper()

        # Verificar se é coluna de total (contém "TOTAL")
        if 'TOTAL' in cell_text:
            col_inicio_totais = col_atual
            logger.debug(f"Coluna de TOTAIS identificada na coluna {col_atual}: '{cell_value}'")
            break

        # Verificar se é coluna de mês (contém "ORÇADO" ou "ORCADO")
        if 'ORÇADO' in cell_text or 'ORCADO' in cell_text:
            # Extrair mês e ano do texto
            # Formato: "JANEIRO 2025\nORÇADO" ou "JANEIRO 2025 ORÇADO"
            linhas = cell_text.split('\n')
            primeira_linha = linhas[0].strip()

            # Tentar extrair mês e ano
            partes = primeira_linha.split()
            mes_nome = None
            ano = None

            for i, parte in enumerate(partes):
                # Procurar mês
                if parte in meses_map and mes_nome is None:
                    mes_nome = parte
                # Procurar ano (número de 4 dígitos)
                elif parte.isdigit() and len(parte) == 4 and ano is None:
                    ano = int(parte)

            if mes_nome and ano:
                mes_numero = meses_map[mes_nome]
                mes_info = {
                    'mes_nome': mes_nome.capitalize(),
                    'mes_numero': mes_numero,
                    'ano': ano,
                    'col_inicio': col_atual
                }
                meses_info.append(mes_info)

                logger.debug(f"Mês identificado: {mes_nome.capitalize()} {ano} (coluna {col_atual})")

                # Pular as próximas 3 colunas deste mês (Realizado, % Atingido, Diferença)
                col_atual += 4
            else:
                logger.warning(f"Não foi possível extrair mês/ano de: '{cell_value}'")
                col_atual += 1
        else:
            col_atual += 1

    # Se não encontrou col_inicio_totais, assume últimas 3 colunas
    if col_inicio_totais is None:
        col_inicio_totais = total_colunas - 2
        logger.warning(f"Coluna de totais não identificada, assumindo coluna {col_inicio_totais}")

    logger.debug(f"Resumo: {len(meses_info)} meses identificados, coluna de totais: {col_inicio_totais}")

    return {
        'meses': meses_info,
        'col_inicio_totais': col_inicio_totais,
        'num_meses': len(meses_info)
    }


def formatar_numero(valor):
    """Formata um número para exibição legível (ou 'N/A' se None)"""
    if valor is None:
        return "N/A"

    if isinstance(valor, (int, float)) and abs(valor) < 0.01:
        return "0.00"

    try:
        return f"{float(valor):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return str(valor)


def processar_item_hierarquico(col_a, row_values, meses_info, col_inicio_totais, linha):
    """
    Processa um item hierárquico com a NOVA estrutura (meses dinâmicos)

    Estrutura DINÂMICA:
    - Coluna 0 (A): Nome/Código
    - Colunas B+: Meses (4 colunas cada: Orçado, Realizado, % Atingido, Diferença)
    - Últimas 3 colunas: Totais (Orçado Total, Realizado Total, Pendente Total)

    Args:
        col_a: Valor da coluna A (código e nome)
        row_values: Lista com todos os valores da linha
        meses_info: Lista com informações dos meses extraídos do cabeçalho
        col_inicio_totais: Coluna onde começam os totais
        linha: Número da linha atual

    Returns:
        dict: Dados estruturados do item
    """
    # Extrair código e nome
    codigo, nome, nivel = extrair_codigo_e_nome(col_a)

    # Processar dados mensais usando as colunas identificadas no cabeçalho
    dados_meses = []

    for mes_info in meses_info:
        col_inicio = mes_info['col_inicio'] - 1  # Converter para índice (col 2 = índice 1)

        # Cada mês tem 4 colunas fixas:
        # 0: Valor Orçado
        # 1: Valor Realizado
        # 2: % Atingido (já vem como número, ex: 86.06 para 86,06%)
        # 3: Valor Diferença

        valor_orcado = converter_valor(row_values[col_inicio]) if col_inicio < len(row_values) else None
        valor_realizado = converter_valor(row_values[col_inicio + 1]) if col_inicio + 1 < len(row_values) else None
        perc_atingido = converter_valor(row_values[col_inicio + 2]) if col_inicio + 2 < len(row_values) else None
        valor_diferenca = converter_valor(row_values[col_inicio + 3]) if col_inicio + 3 < len(row_values) else None

        mes_data = {
            'mes_numero': mes_info['mes_numero'],
            'mes_nome': mes_info['mes_nome'],
            'ano': mes_info['ano'],
            'valor_orcado': valor_orcado,
            'valor_realizado': valor_realizado,
            'perc_atingido': perc_atingido,
            'valor_diferenca': valor_diferenca,
        }
        dados_meses.append(mes_data)

    # Processar resultados totais (coluna identificada no cabeçalho)
    idx_resultados_inicio = col_inicio_totais - 1  # Converter para índice

    valor_orcado_total = converter_valor(row_values[idx_resultados_inicio]) if idx_resultados_inicio < len(row_values) else None
    valor_realizado_total = converter_valor(row_values[idx_resultados_inicio + 1]) if idx_resultados_inicio + 1 < len(row_values) else None
    valor_pendente_total = converter_valor(row_values[idx_resultados_inicio + 2]) if idx_resultados_inicio + 2 < len(row_values) else None

    resultados = {
        'valor_orcado_total': valor_orcado_total,
        'valor_realizado_total': valor_realizado_total,
        'valor_pendente_total': valor_pendente_total
    }

    return {
        'codigo': codigo,
        'nome': nome,
        'nivel_hierarquia': nivel,
        'linha': linha,
        'dados_mensais': dados_meses,
        'resultados_totais': resultados
    }


def calcular_totais_fluxo_caixa(itens_hierarquicos, meses_info):
    """
    Calcula os totais do 1º cenário: RESULTADO POR FLUXO DE CAIXA

    Para cada mês, calcula 3 totais em 4 colunas:
    - Orçamento: Receita, Despesa, Geral
    - Realizado: Receita, Despesa, Geral
    - % Atingido: Receita%, Despesa%, Geral%
    - Diferença: Receita, Despesa, Geral

    Args:
        itens_hierarquicos: Lista de itens processados
        meses_info: Lista com informações dos meses extraídos do cabeçalho

    Returns:
        dict: Totais calculados por mês
    """

    # Estrutura de retorno
    totais = {
        'fluxo_caixa': {},  # Será preenchido por mês (chave: 'ano_mes')
        'real': {},         # Para depois (vazio por enquanto)
        'real_mp': {}       # Para depois (vazio por enquanto)
    }

    # Encontrar os itens "1 - RECEITA" e "2 - DESPESAS"
    item_receita = None
    item_despesa = None

    for item in itens_hierarquicos:
        nome_upper = item['nome'].upper()
        codigo = item['codigo']

        # Procurar por "1 - RECEITA" ou similar
        if codigo == "1" or ("RECEITA" in nome_upper and codigo.startswith("1")):
            if not item_receita:  # Pegar o primeiro
                item_receita = item
                logger.debug(f"Item RECEITA encontrado: [{item['codigo']}] {item['nome']}")

        # Procurar por "2 - DESPESAS" ou similar
        if codigo == "2" or ("DESPESA" in nome_upper and codigo.startswith("2")):
            if not item_despesa:  # Pegar o primeiro
                item_despesa = item
                logger.debug(f"Item DESPESA encontrado: [{item['codigo']}] {item['nome']}")

    if not item_receita or not item_despesa:
        logger.warning("ATENÇÃO: Não foi possível encontrar itens RECEITA e/ou DESPESAS!")
        logger.warning(f"Item RECEITA: {'Encontrado' if item_receita else 'NÃO ENCONTRADO'}")
        logger.warning(f"Item DESPESA: {'Encontrado' if item_despesa else 'NÃO ENCONTRADO'}")
        return totais

    # Calcular para cada mês encontrado na planilha
    for mes_info in meses_info:
        mes_numero = mes_info['mes_numero']
        ano = mes_info['ano']
        chave_mes = f"{ano}_{mes_numero}"  # Ex: "2025_3" para Março 2025

        # Encontrar dados deste mês
        dados_mes_receita = next((m for m in item_receita['dados_mensais'] if m['mes_numero'] == mes_numero and m['ano'] == ano), None)
        dados_mes_despesa = next((m for m in item_despesa['dados_mensais'] if m['mes_numero'] == mes_numero and m['ano'] == ano), None)

        if not dados_mes_receita or not dados_mes_despesa:
            logger.warning(f"{mes_info['mes_nome']} {ano}: Dados não encontrados")
            continue

        # ====================================================================
        # COLUNA 1 - ORÇAMENTO
        # ====================================================================
        orcamento_receita = dados_mes_receita['valor_orcado'] or 0
        orcamento_despesa = dados_mes_despesa['valor_orcado'] or 0
        orcamento_geral = orcamento_receita - orcamento_despesa

        # ====================================================================
        # COLUNA 2 - REALIZADO
        # ====================================================================
        realizado_receita = dados_mes_receita['valor_realizado'] or 0
        realizado_despesa = dados_mes_despesa['valor_realizado'] or 0
        realizado_geral = realizado_receita - realizado_despesa

        # ====================================================================
        # COLUNA 3 - % ATINGIDO
        # ====================================================================
        # Fórmula: (Realizado / Orçado) * 100
        if orcamento_receita != 0:
            perc_receita = (realizado_receita / orcamento_receita) * 100
        else:
            perc_receita = 0

        if orcamento_despesa != 0:
            perc_despesa = (realizado_despesa / orcamento_despesa) * 100
        else:
            perc_despesa = 0

        if orcamento_geral != 0:
            perc_geral = (realizado_geral / orcamento_geral) * 100
        else:
            perc_geral = 0

        # ====================================================================
        # COLUNA 4 - DIFERENÇA
        # ====================================================================
        # Receita: Realizado - Orçado
        # Despesa: Orçado - Realizado (INVERTIDO!)
        # Geral: Receita + Despesa
        diferenca_receita = realizado_receita - orcamento_receita
        diferenca_despesa = orcamento_despesa - realizado_despesa  # INVERTIDO
        diferenca_geral = diferenca_receita + diferenca_despesa

        # Salvar totais deste mês usando chave ano_mes
        totais['fluxo_caixa'][chave_mes] = {
            'mes_numero': mes_numero,
            'ano': ano,
            'mes_nome': mes_info['mes_nome'],
            'orcamento': {
                'receita': orcamento_receita,
                'despesa': orcamento_despesa,
                'geral': orcamento_geral
            },
            'realizado': {
                'receita': realizado_receita,
                'despesa': realizado_despesa,
                'geral': realizado_geral
            },
            'perc_atingido': {
                'receita': perc_receita,
                'despesa': perc_despesa,
                'geral': perc_geral
            },
            'diferenca': {
                'receita': diferenca_receita,
                'despesa': diferenca_despesa,
                'geral': diferenca_geral
            }
        }

        # Log apenas do primeiro mês para não poluir
        if len(totais['fluxo_caixa']) == 1:
            logger.debug(f"{mes_info['mes_nome']} {ano} - Exemplo de cálculo do Fluxo de Caixa")
            logger.debug(f"ORÇAMENTO → Receita: R$ {formatar_numero(orcamento_receita)} | Despesa: R$ {formatar_numero(orcamento_despesa)} | Geral: R$ {formatar_numero(orcamento_geral)}")

    logger.debug(f"Totais do Fluxo de Caixa calculados para {len(totais['fluxo_caixa'])} meses")

    # ========================================================================
    # CALCULAR 2º CENÁRIO: RESULTADO REAL
    # ========================================================================
    logger.debug("CALCULANDO TOTAIS - RESULTADO REAL")

    # Itens a subtrair da RECEITA para o Resultado Real
    itens_subtrair_receita = ['RECEITA EMPRESTIMO', 'OUTRAS RECEITAS']

    # Itens a subtrair da DESPESA para o Resultado Real
    itens_subtrair_despesa = [
        'OUTRAS DESPESAS NÃO DEDUTIVEIS',
        'Distribuição de lucro Associados',
        'SAIDA- EMPRESTIMOS',
        'INVESTIMENTOS'
    ]

    # Função auxiliar para buscar valores de um item por nome
    def buscar_valores_item(nome_item):
        """Retorna os dados_mensais de um item pelo nome"""
        for item in itens_hierarquicos:
            if item['nome'].strip().upper() == nome_item.upper():
                return item.get('dados_mensais', [])
        return []

    # Função auxiliar para calcular total de subtração
    def calcular_total_subtracao(nomes_itens, mes_num, campo):
        """
        Soma os valores de múltiplos itens para um mês específico
        campo: 'valor_orcado', 'valor_realizado'
        """
        total = 0
        for nome in nomes_itens:
            dados_mensais = buscar_valores_item(nome)
            for mes_data in dados_mensais:
                if mes_data['mes_numero'] == mes_num:
                    valor = mes_data.get(campo, 0)
                    total += valor if valor else 0
                    break
        return total

    # Calcular para cada mês
    for mes_info in meses_info:
        mes_numero = mes_info['mes_numero']
        ano = mes_info['ano']
        chave_mes = f"{ano}_{mes_numero}"

        # Pegar dados do Fluxo de Caixa deste mês
        dados_fc = totais['fluxo_caixa'].get(chave_mes)

        if not dados_fc:
            logger.warning(f"{mes_info['mes_nome']} {ano}: Dados do Fluxo de Caixa não encontrados")
            continue

        # ====================================================================
        # COLUNA 1 - ORÇAMENTO
        # ====================================================================
        orcamento_receita_fc = dados_fc['orcamento']['receita']
        orcamento_despesa_fc = dados_fc['orcamento']['despesa']

        # Subtrair os itens específicos
        subtracao_receita_orcado = calcular_total_subtracao(itens_subtrair_receita, mes_numero, 'valor_orcado')
        subtracao_despesa_orcado = calcular_total_subtracao(itens_subtrair_despesa, mes_numero, 'valor_orcado')

        orcamento_receita_real = orcamento_receita_fc - subtracao_receita_orcado
        orcamento_despesa_real = orcamento_despesa_fc - subtracao_despesa_orcado
        orcamento_geral_real = orcamento_receita_real - orcamento_despesa_real

        # ====================================================================
        # COLUNA 2 - REALIZADO
        # ====================================================================
        realizado_receita_fc = dados_fc['realizado']['receita']
        realizado_despesa_fc = dados_fc['realizado']['despesa']

        # Subtrair os itens específicos
        subtracao_receita_realizado = calcular_total_subtracao(itens_subtrair_receita, mes_numero, 'valor_realizado')
        subtracao_despesa_realizado = calcular_total_subtracao(itens_subtrair_despesa, mes_numero, 'valor_realizado')

        realizado_receita_real = realizado_receita_fc - subtracao_receita_realizado
        realizado_despesa_real = realizado_despesa_fc - subtracao_despesa_realizado
        realizado_geral_real = realizado_receita_real - realizado_despesa_real

        # ====================================================================
        # COLUNA 3 - % ATINGIDO
        # ====================================================================
        # Fórmula: (Realizado / Orçado) * 100
        if orcamento_receita_real != 0:
            perc_receita_real = (realizado_receita_real / orcamento_receita_real) * 100
        else:
            perc_receita_real = 0

        if orcamento_despesa_real != 0:
            perc_despesa_real = (realizado_despesa_real / orcamento_despesa_real) * 100
        else:
            perc_despesa_real = 0

        if orcamento_geral_real != 0:
            perc_geral_real = (realizado_geral_real / orcamento_geral_real) * 100
        else:
            perc_geral_real = 0

        # ====================================================================
        # COLUNA 4 - DIFERENÇA
        # ====================================================================
        # Receita: Realizado - Orçado
        # Despesa: Orçado - Realizado (INVERTIDO!)
        # Geral: Receita + Despesa
        diferenca_receita_real = realizado_receita_real - orcamento_receita_real
        diferenca_despesa_real = orcamento_despesa_real - realizado_despesa_real  # INVERTIDO
        diferenca_geral_real = diferenca_receita_real + diferenca_despesa_real

        # Salvar totais deste mês usando chave ano_mes
        totais['real'][chave_mes] = {
            'mes_numero': mes_numero,
            'ano': ano,
            'mes_nome': mes_info['mes_nome'],
            'orcamento': {
                'receita': orcamento_receita_real,
                'despesa': orcamento_despesa_real,
                'geral': orcamento_geral_real
            },
            'realizado': {
                'receita': realizado_receita_real,
                'despesa': realizado_despesa_real,
                'geral': realizado_geral_real
            },
            'perc_atingido': {
                'receita': perc_receita_real,
                'despesa': perc_despesa_real,
                'geral': perc_geral_real
            },
            'diferenca': {
                'receita': diferenca_receita_real,
                'despesa': diferenca_despesa_real,
                'geral': diferenca_geral_real
            }
        }

        # Log apenas do primeiro mês
        if len(totais['real']) == 1:
            logger.debug(f"{mes_info['mes_nome']} {ano} - Exemplo de cálculo (Resultado Real)")

    logger.debug(f"Totais do Resultado Real calculados para {len(totais['real'])} meses")

    # ========================================================================
    # CALCULAR 3º CENÁRIO: RESULTADO REAL + CUSTO MATÉRIA PRIMA
    # ========================================================================
    logger.debug("CALCULANDO TOTAIS - RESULTADO REAL + CUSTO MATÉRIA PRIMA")

    # Item a buscar: CUSTO MATERIA PRIMA
    item_custo_mp_nome = 'CUSTO MATERIA PRIMA'

    # Calcular para cada mês
    for mes_info in meses_info:
        mes_numero = mes_info['mes_numero']
        ano = mes_info['ano']
        chave_mes = f"{ano}_{mes_numero}"

        # Pegar dados do cenário "Real" deste mês
        dados_real = totais['real'].get(chave_mes)

        if not dados_real:
            logger.warning(f"{mes_info['mes_nome']} {ano}: Dados do Resultado Real não encontrados")
            continue

        # ====================================================================
        # COLUNA 1 - ORÇAMENTO
        # ====================================================================
        # Mesmos valores do cenário "Real"
        orcamento_receita_mp = dados_real['orcamento']['receita']
        orcamento_despesa_mp = dados_real['orcamento']['despesa']
        orcamento_geral_mp = dados_real['orcamento']['geral']

        # ====================================================================
        # COLUNA 2 - REALIZADO
        # ====================================================================
        # Total Receita = igual ao cenário "Real"
        realizado_receita_mp = dados_real['realizado']['receita']

        # Buscar valores de CUSTO MATERIA PRIMA
        dados_custo_mp = buscar_valores_item(item_custo_mp_nome)
        custo_mp_orcado = 0
        custo_mp_realizado = 0

        if dados_custo_mp:
            for mes_data in dados_custo_mp:
                if mes_data['mes_numero'] == mes_numero and mes_data['ano'] == ano:
                    custo_mp_orcado = mes_data.get('valor_orcado', 0) or 0
                    custo_mp_realizado = mes_data.get('valor_realizado', 0) or 0
                    break

        # Total Despesa (fórmula complexa):
        # ((CUSTO_MP_ORCADO / RECEITA_ORCADO) * RECEITA_REALIZADO - CUSTO_MP_REALIZADO)
        # + DESPESA_REAL_REALIZADO - CUSTO_MP_REALIZADO

        despesa_real_realizado = dados_real['realizado']['despesa']

        if orcamento_receita_mp != 0:
            parte1 = ((custo_mp_orcado / orcamento_receita_mp) * realizado_receita_mp) - custo_mp_realizado
        else:
            parte1 = 0

        realizado_despesa_mp = parte1 + despesa_real_realizado - custo_mp_realizado
        realizado_geral_mp = realizado_receita_mp - realizado_despesa_mp

        # ====================================================================
        # COLUNA 3 - % ATINGIDO
        # ====================================================================
        # Fórmula: (Realizado / Orçado) * 100
        if orcamento_receita_mp != 0:
            perc_receita_mp = (realizado_receita_mp / orcamento_receita_mp) * 100
        else:
            perc_receita_mp = 0

        if orcamento_despesa_mp != 0:
            perc_despesa_mp = (realizado_despesa_mp / orcamento_despesa_mp) * 100
        else:
            perc_despesa_mp = 0

        if orcamento_geral_mp != 0:
            perc_geral_mp = (realizado_geral_mp / orcamento_geral_mp) * 100
        else:
            perc_geral_mp = 0

        # ====================================================================
        # COLUNA 4 - DIFERENÇA
        # ====================================================================
        # Receita: Realizado - Orçado
        # Despesa: Orçado - Realizado (INVERTIDO!)
        # Geral: Receita + Despesa
        diferenca_receita_mp = realizado_receita_mp - orcamento_receita_mp
        diferenca_despesa_mp = orcamento_despesa_mp - realizado_despesa_mp  # INVERTIDO
        diferenca_geral_mp = diferenca_receita_mp + diferenca_despesa_mp

        # Salvar totais deste mês usando chave ano_mes
        totais['real_mp'][chave_mes] = {
            'mes_numero': mes_numero,
            'ano': ano,
            'mes_nome': mes_info['mes_nome'],
            'orcamento': {
                'receita': orcamento_receita_mp,
                'despesa': orcamento_despesa_mp,
                'geral': orcamento_geral_mp
            },
            'realizado': {
                'receita': realizado_receita_mp,
                'despesa': realizado_despesa_mp,
                'geral': realizado_geral_mp
            },
            'perc_atingido': {
                'receita': perc_receita_mp,
                'despesa': perc_despesa_mp,
                'geral': perc_geral_mp
            },
            'diferenca': {
                'receita': diferenca_receita_mp,
                'despesa': diferenca_despesa_mp,
                'geral': diferenca_geral_mp
            }
        }

        # Log apenas do primeiro mês
        if len(totais['real_mp']) == 1:
            logger.debug(f"{mes_info['mes_nome']} {ano} - Exemplo de cálculo (Resultado Real + Custo MP)")

    logger.debug(f"Totais do Resultado Real + Custo MP calculados para {len(totais['real_mp'])} meses")

    return totais


# ============================================================================
# FUNÇÃO PRINCIPAL DE PROCESSAMENTO
# ============================================================================

def process_bpo_file(file):
    """
    Processa arquivo Excel de BPO Financeiro (NOVA ESTRUTURA) e retorna dados estruturados.

    Estrutura da planilha:
    - Sheet: "Sheet"
    - Linha 1: Cabeçalho
    - Linha 2+: Dados começam
    - Coluna A: Código hierárquico e nome (ex: "1.01 - RECEITA VENDA SERVIÇO")
    - Coluna B+: Dados mensais (4 colunas por mês: Orçado, Realizado, % Ating, Diferença)
    - Últimas 3 colunas: Totais (Orçado Total, Realizado Total, Pendente Total)

    Args:
        file: Arquivo Excel (.xlsx ou .xls)

    Returns:
        dict: {
            'itens_hierarquicos': [...],  # Itens com hierarquia
            'totais_calculados': {},      # Para adicionar depois (quando souber a fórmula)
            'metadados': {...}            # Info sobre meses, totais, etc
        }
    """

    try:
        logger.info("PROCESSANDO PLANILHA BPO (NOVA ESTRUTURA)")

        # Carregar workbook (data_only=True para pegar valores calculados ao invés de fórmulas)
        wb = load_workbook(file, data_only=True)

        # Selecionar sheet 'Sheet'
        sheet_name = 'Sheet'
        if sheet_name not in wb.sheetnames:
            raise Exception(f"Sheet '{sheet_name}' não encontrada. Sheets disponíveis: {wb.sheetnames}")

        sheet = wb[sheet_name]
        logger.debug(f"Sheet '{sheet_name}' encontrada")

        # Identificar estrutura da planilha
        total_colunas = sheet.max_column
        logger.debug(f"Total de colunas na planilha: {total_colunas}")

        # Ler o cabeçalho para extrair informações dos meses dinamicamente
        info_cabecalho = extrair_meses_do_cabecalho(sheet)
        meses_info = info_cabecalho['meses']
        col_inicio_totais = info_cabecalho['col_inicio_totais']
        num_meses = info_cabecalho['num_meses']

        meses_str = ', '.join([f"{m['mes_nome']} {m['ano']}" for m in meses_info])
        logger.info(f"Número de meses detectados: {num_meses} ({meses_str})")

        # Processar itens hierárquicos (LINHA 2 em diante)
        itens_hierarquicos = []
        linha_atual = 2  # Começa na linha 2 (linha 1 = cabeçalho)

        logger.debug("PROCESSANDO ITENS HIERÁRQUICOS")

        while True:
            row_values = []
            for col in range(1, total_colunas + 1):
                cell_value = sheet.cell(row=linha_atual, column=col).value
                row_values.append(cell_value)

            # Verifica se linha está completamente vazia (fim da planilha)
            if all(v is None or str(v).strip() == '' for v in row_values):
                logger.debug(f"Linha {linha_atual}: Vazia - fim dos dados")
                break

            # Processar item se coluna A tem conteúdo
            col_a = row_values[0]
            if col_a and str(col_a).strip():
                item = processar_item_hierarquico(
                    col_a,
                    row_values,
                    meses_info,
                    col_inicio_totais,
                    linha_atual
                )
                itens_hierarquicos.append(item)

            linha_atual += 1

        logger.debug(f"Total de itens processados: {len(itens_hierarquicos)}")

        # ========================================================================
        # CALCULAR TOTAIS (1º CENÁRIO: RESULTADO POR FLUXO DE CAIXA)
        # ========================================================================
        logger.debug("CALCULANDO TOTAIS - RESULTADO POR FLUXO DE CAIXA")

        totais_calculados = calcular_totais_fluxo_caixa(itens_hierarquicos, meses_info)

        # Montar estrutura final
        dados_processados = {
            'itens_hierarquicos': itens_hierarquicos,
            'totais_calculados': totais_calculados,
            'metadados': {
                'total_colunas': total_colunas,
                'num_meses': num_meses,
                'meses_info': meses_info,
                'total_itens': len(itens_hierarquicos)
            }
        }

        logger.info(f"PROCESSAMENTO CONCLUÍDO COM SUCESSO! Itens: {len(itens_hierarquicos)}, Meses: {num_meses}, Colunas: {total_colunas}")

        return dados_processados

    except Exception as e:
        logger.error(f"ERRO AO PROCESSAR ARQUIVO BPO: {str(e)}")
        import traceback
        traceback.print_exc()
        raise Exception(f"Erro no processamento do BPO: {str(e)}")


def validate_bpo_data(dados):
    """
    Valida os dados de BPO processados antes de salvar no banco.

    Args:
        dados (dict): Dados processados pela função process_bpo_file()

    Returns:
        tuple: (bool, str) - (True/False, mensagem de erro/sucesso)
    """
    if not dados:
        return False, "Dados vazios ou inválidos"

    if 'itens_hierarquicos' not in dados:
        return False, "Estrutura de dados inválida: falta campo 'itens_hierarquicos'"

    if 'metadados' not in dados:
        return False, "Estrutura de dados inválida: falta campo 'metadados'"

    if len(dados['itens_hierarquicos']) == 0:
        return False, "Nenhum item hierárquico encontrado"

    return True, "Validação OK"
