from flask import Blueprint, render_template, session, redirect, url_for, request, flash, jsonify
from datetime import datetime
from utils.logger import get_logger

# Inicializar logger
logger = get_logger('user_pages')

user_bp = Blueprint('user', __name__)

@user_bp.route('/user/selecionar-empresa')
def selecionar_empresa():
    """Página para o usuário selecionar qual empresa deseja acessar"""
    if 'user_email' in session and session.get('user_role') == 'user':
        from models.user_manager import UserManager

        user_manager = UserManager()
        user_data = user_manager.find_user_by_email(session.get('user_email'))

        if not user_data:
            flash("Erro ao carregar dados do usuário.", "danger")
            return redirect(url_for('index.login'))

        # Busca empresas vinculadas ao usuário
        empresas = user_manager.get_empresas_do_usuario(user_data['id'])
        user_manager.close()

        # Se não tiver empresa vinculada
        if not empresas:
            flash("Você não está vinculado a nenhuma empresa. Entre em contato com o administrador.", "warning")
            return redirect(url_for('user.logout'))

        # Se tiver apenas 1 empresa, redireciona direto pro dashboard
        if len(empresas) == 1:
            session['empresa_id'] = empresas[0]['id']
            session['empresa_nome'] = empresas[0]['nome']
            return redirect(url_for('user.user_dashboard'))

        # Se tiver múltiplas empresas, exibe página de seleção
        return render_template(
            'user/selecionar_empresa.html',
            user=user_data,
            empresas=empresas
        )
    else:
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))


@user_bp.route('/user/definir-empresa/<int:empresa_id>')
def definir_empresa(empresa_id):
    """Define a empresa que o usuário quer acessar"""
    if 'user_email' in session and session.get('user_role') == 'user':
        from models.user_manager import UserManager
        from models.company_manager import CompanyManager

        user_manager = UserManager()
        user_data = user_manager.find_user_by_email(session.get('user_email'))

        if not user_data:
            flash("Erro ao carregar dados do usuário.", "danger")
            return redirect(url_for('index.login'))

        # Verifica se o usuário tem acesso a essa empresa
        empresas = user_manager.get_empresas_do_usuario(user_data['id'])
        user_manager.close()

        empresa_encontrada = None
        for empresa in empresas:
            if empresa['id'] == empresa_id:
                empresa_encontrada = empresa
                break

        if not empresa_encontrada:
            flash("Você não tem acesso a esta empresa.", "danger")
            return redirect(url_for('user.selecionar_empresa'))

        # Salva na sessão
        session['empresa_id'] = empresa_id
        session['empresa_nome'] = empresa_encontrada['nome']

        flash(f"Empresa '{empresa_encontrada['nome']}' selecionada com sucesso!", "success")
        return redirect(url_for('user.user_dashboard'))
    else:
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))


@user_bp.route('/user/dashboard')
def user_dashboard():
    """Dashboard principal do usuário - mostra dados anuais da empresa selecionada"""
    if 'user_email' in session and session.get('user_role') == 'user':
        from models.user_manager import UserManager
        from models.company_manager import CompanyManager

        # Verifica se tem empresa selecionada
        if 'empresa_id' not in session:
            return redirect(url_for('user.selecionar_empresa'))

        # Pega informações do usuário logado
        user_manager = UserManager()
        user_data = user_manager.find_user_by_email(session.get('user_email'))
        user_manager.close()

        if not user_data:
            flash("Erro ao carregar dados do usuário.", "danger")
            return redirect(url_for('index.login'))

        empresa_id = session.get('empresa_id')
        empresa_nome = session.get('empresa_nome')

        # Busca anos com dados disponíveis para esta empresa
        company_manager = CompanyManager()
        anos_disponiveis = company_manager.get_anos_com_dados(empresa_id)
        company_manager.close()

        # Renderizar dashboard de viabilidade do usuário
        return render_template(
            'user/dashboard_empresa.html',
            user=user_data,
            empresa_nome=empresa_nome,
            empresa_id=empresa_id,
            anos_disponiveis=anos_disponiveis,
            is_user_view=True  # Flag para o template saber que é visualização de usuário
        )
    else:
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))


@user_bp.route('/user/api/dados-bpo-tabela/<int:empresa_id>')
def api_dados_bpo_tabela(empresa_id):
    """
    API para retornar dados BPO em formato tabular (tipo planilha)
    Filtra por período e retorna itens hierárquicos com dados mensais
    """
    if 'user_email' not in session or session.get('user_role') != 'user':
        return jsonify({'error': 'Não autorizado'}), 403

    # Verificar se usuário tem acesso a esta empresa
    empresa_id_session = session.get('empresa_id')
    if empresa_id != empresa_id_session:
        return jsonify({'error': 'Acesso negado a esta empresa'}), 403

    try:
        from models.company_manager import CompanyManager
        import json

        # Pegar parâmetros do filtro
        mes_inicio = int(request.args.get('mes_inicio', 1))
        ano_inicio = int(request.args.get('ano_inicio', 2025))
        mes_fim = int(request.args.get('mes_fim', 12))
        ano_fim = int(request.args.get('ano_fim', 2025))

        company_manager = CompanyManager()

        # Buscar todos os meses do período
        meses_dados = []
        ano_atual = ano_inicio
        mes_atual = mes_inicio

        while (ano_atual < ano_fim) or (ano_atual == ano_fim and mes_atual <= mes_fim):
            dados = company_manager.buscar_dados_bpo_empresa(empresa_id, ano_atual, mes_atual)
            if dados:
                dados_json = dados['dados']  # Já vem parseado
                meses_dados.append({
                    'mes': mes_atual,
                    'ano': ano_atual,
                    'dados': dados_json
                })

            # Avançar para próximo mês
            mes_atual += 1
            if mes_atual > 12:
                mes_atual = 1
                ano_atual += 1

        company_manager.close()

        if not meses_dados:
            return jsonify({'error': 'Nenhum dado encontrado para o período'}), 404

        # Consolidar itens hierárquicos de todos os meses
        itens_consolidados = {}
        meses_info = []

        for mes_data in meses_dados:
            mes = mes_data['mes']
            ano = mes_data['ano']
            dados = mes_data['dados']

            # Adicionar info do mês
            meses_nomes = ['', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                          'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
            meses_info.append({
                'mes_numero': mes,
                'mes_nome': meses_nomes[mes],
                'ano': ano
            })

            # Processar cada item hierárquico
            for item in dados.get('itens_hierarquicos', []):
                codigo = item['codigo']

                if codigo not in itens_consolidados:
                    itens_consolidados[codigo] = {
                        'codigo': codigo,
                        'nome': item['nome'],
                        'nivel_hierarquia': item['nivel_hierarquia'],
                        'dados_mensais': {}
                    }

                # Adicionar dados deste mês
                if item.get('dados_mensais') and len(item['dados_mensais']) > 0:
                    dados_mes = item['dados_mensais'][0]  # Pega primeiro mês (deveria ser único)
                    chave_mes = f"{ano}_{mes}"
                    itens_consolidados[codigo]['dados_mensais'][chave_mes] = {
                        'mes_numero': mes,
                        'ano': ano,
                        'valor_orcado': dados_mes.get('valor_orcado'),
                        'valor_realizado': dados_mes.get('valor_realizado'),
                        'perc_atingido': dados_mes.get('perc_atingido'),
                        'valor_diferenca': dados_mes.get('valor_diferenca')
                    }

        # Converter para lista ordenada por código
        itens_lista = sorted(itens_consolidados.values(), key=lambda x: x['codigo'])

        # Buscar totais calculados para o período (se existirem)
        totais_calculados = {
            'fluxo_caixa': {},
            'real': {},
            'real_mp': {}
        }

        for mes_data in meses_dados:
            mes = mes_data['mes']
            ano = mes_data['ano']
            dados = mes_data['dados']
            chave_mes = f"{ano}_{mes}"

            # Extrair totais calculados
            totais = dados.get('totais_calculados', {})

            # Os totais_calculados vêm com TODOS os meses do Excel
            # Então apenas precisamos mesclar os dados de cada cenário
            for cenario in ['fluxo_caixa', 'real', 'real_mp']:
                if cenario in totais and isinstance(totais[cenario], dict):
                    # Mesclar todos os meses deste cenário
                    for mes_key, mes_value in totais[cenario].items():
                        # Normalizar chave: converter formato antigo '1' para novo '2025_1'
                        # Se mes_key é apenas número (formato antigo), converter
                        if mes_key.isdigit():
                            # Formato antigo: apenas número do mês
                            mes_num = int(mes_key)
                            # Pegar o ano dos metadados do mes_value ou usar ano atual
                            ano_mes = mes_value.get('ano', ano)
                            chave_normalizada = f"{ano_mes}_{mes_num}"
                        else:
                            # Formato novo: já está como 'ano_mes'
                            chave_normalizada = mes_key

                        # Adicionar apenas se ainda não tiver (evitar duplicação)
                        if chave_normalizada not in totais_calculados[cenario]:
                            totais_calculados[cenario][chave_normalizada] = mes_value

        return jsonify({
            'itens': itens_lista,
            'meses': meses_info,
            'totais_calculados': totais_calculados
        })

    except Exception as e:
        logger.error(f"Erro na API dados-bpo-tabela: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@user_bp.route('/user/relatorio-pdf/<int:empresa_id>/<int:ano>/<grupo_viabilidade>')
def gerar_relatorio_pdf(empresa_id, ano, grupo_viabilidade):
    """Gera PDF do relatório de viabilidade (rota para usuários)"""
    if 'user_email' not in session or session.get('user_role') != 'user':
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))

    # Verificar se usuário tem acesso a esta empresa
    empresa_id_session = session.get('empresa_id')
    if empresa_id != empresa_id_session:
        flash("Acesso negado a esta empresa.", "danger")
        return redirect(url_for('user.user_dashboard'))

    try:
        from models.company_manager import CompanyManager
        from xhtml2pdf import pisa
        import io

        # Buscar empresa
        company_manager = CompanyManager()
        empresa = company_manager.buscar_empresa_por_id(empresa_id)

        if not empresa:
            flash("Empresa não encontrada.", "danger")
            company_manager.close()
            return redirect(url_for('user.user_dashboard'))

        # Buscar template do relatório
        template_data = company_manager.buscar_template_relatorio(empresa_id, ano)
        company_manager.close()

        if not template_data:
            flash(f"Template de relatório não encontrado para o ano {ano}.", "warning")
            return redirect(url_for('user.user_dashboard'))

        # Pegar o texto do template (já vem pronto com os valores)
        conteudo_texto = template_data['template']

        # Adicionar CSS para formatação do PDF
        css_style = """
        <style>
            @page { size: A4; margin: 2cm; }
            body { font-family: Arial, sans-serif; font-size: 11pt; line-height: 1.6; color: #333; }
            h1 { color: #2c3e50; font-size: 18pt; margin-top: 1em; }
            h2 { color: #34495e; font-size: 14pt; margin-top: 1em; }
            h3 { color: #34495e; font-size: 12pt; margin-top: 0.8em; }
            table { width: 100%; border-collapse: collapse; margin: 1em 0; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #f2f2f2; font-weight: bold; }
            .header { text-align: center; margin-bottom: 2em; border-bottom: 2px solid #2c3e50; padding-bottom: 1em; }
            .section { margin: 1.5em 0; }
            p { margin: 0.5em 0; }
        </style>
        """

        # Converter quebras de linha do texto em tags HTML <br>
        conteudo_html = conteudo_texto.replace('\n', '<br>\n')

        html_completo = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            {css_style}
        </head>
        <body>
            <div class="header">
                <h1>RELATÓRIO DE VIABILIDADE FINANCEIRA</h1>
                <p><strong>{empresa['nome']}</strong> - Ano: {ano}</p>
                <p>Cenário: {grupo_viabilidade}</p>
            </div>
            <div class="content">
                {conteudo_html}
            </div>
        </body>
        </html>
        """

        # Gerar PDF usando xhtml2pdf (funciona melhor no Windows)
        pdf_file = io.BytesIO()
        pisa_status = pisa.CreatePDF(
            html_completo.encode('utf-8'),
            dest=pdf_file,
            encoding='utf-8'
        )

        if pisa_status.err:
            raise Exception(f"Erro ao gerar PDF: {pisa_status.err}")

        pdf_file.seek(0)

        # Retornar PDF como download
        from flask import send_file
        return send_file(
            pdf_file,
            mimetype='application/pdf',
            as_attachment=True,
            download_name=f'Relatorio_Viabilidade_{empresa["nome"]}_{ano}_{grupo_viabilidade}.pdf'
        )

    except Exception as e:
        logger.error(f"Erro ao gerar PDF: {e}")
        import traceback
        traceback.print_exc()
        flash(f"Erro ao gerar PDF: {str(e)}", "danger")
        return redirect(url_for('user.user_dashboard'))


@user_bp.route('/user/api/relatorio-ia-viabilidade/<int:empresa_id>', methods=['POST'])
def api_relatorio_ia_viabilidade_user(empresa_id):
    """Gera relatório de viabilidade usando IA (Gemini) - rota para usuários"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        return jsonify({"error": "Não autorizado"}), 403

    try:
        from controllers.AI.gemini_utils import gerar_relatorio_viabilidade
        from models.company_manager import CompanyManager

        # Recebe os dados do frontend
        dados = request.get_json()

        if not dados:
            return jsonify({"error": "Dados não fornecidos"}), 400

        # Buscar nome da empresa
        company_manager = CompanyManager()
        empresa = company_manager.buscar_empresa_por_id(empresa_id)
        company_manager.close()

        if not empresa:
            return jsonify({"error": "Empresa não encontrada"}), 404

        # Adiciona o nome da empresa aos dados
        dados['empresa_nome'] = empresa['nome']

        # Gera o relatório usando o Gemini
        relatorio_markdown = gerar_relatorio_viabilidade(dados)

        return jsonify({
            "success": True,
            "relatorio": relatorio_markdown
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório IA: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@user_bp.route('/user/api/relatorio-ia-bpo/<int:empresa_id>', methods=['POST'])
def api_relatorio_ia_bpo_user(empresa_id):
    """Gera relatório executivo de DRE e Performance usando IA (Gemini) - rota para usuários"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        return jsonify({"error": "Não autorizado"}), 403

    try:
        from controllers.AI.gemini_utils import gerar_relatorio_bpo
        from models.company_manager import CompanyManager

        # Recebe os dados do frontend
        dados = request.get_json()

        if not dados:
            return jsonify({"error": "Dados não fornecidos"}), 400

        # Buscar nome da empresa
        company_manager = CompanyManager()
        empresa = company_manager.buscar_empresa_por_id(empresa_id)
        company_manager.close()

        if not empresa:
            return jsonify({"error": "Empresa não encontrada"}), 404

        # Adiciona o nome da empresa aos dados
        dados['empresa_nome'] = empresa['nome']

        # Gera o relatório usando o Gemini
        relatorio_html = gerar_relatorio_bpo(dados)

        return jsonify({
            "success": True,
            "relatorio": relatorio_html
        })

    except Exception as e:
        logger.error(f"Erro ao gerar relatório IA BPO: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@user_bp.route('/user/dados')
def visualizar_dados():
    """Página de visualização detalhada dos dados anuais da empresa selecionada"""
    if 'user_email' in session and session.get('user_role') == 'user':
        from models.user_manager import UserManager
        from models.company_manager import CompanyManager

        # Verifica se tem empresa selecionada
        if 'empresa_id' not in session:
            return redirect(url_for('user.selecionar_empresa'))

        # Pega informações do usuário logado
        user_manager = UserManager()
        user_data = user_manager.find_user_by_email(session.get('user_email'))
        user_manager.close()

        if not user_data:
            flash("Erro ao carregar dados do usuário.", "danger")
            return redirect(url_for('index.login'))

        empresa_id = session.get('empresa_id')
        empresa_nome = session.get('empresa_nome')

        # Parâmetro de filtro (apenas ano)
        ano_selecionado = request.args.get('ano', type=int)

        data_results = None

        if ano_selecionado:
            # Busca dados específicos do ano para esta empresa
            company_manager = CompanyManager()
            data_results = company_manager.buscar_dados_empresa(
                empresa_id,
                ano_selecionado
            )
            company_manager.close()

        return render_template(
            'user/dados.html',
            user=user_data,
            empresa_nome=empresa_nome,
            data_results=data_results,
            ano_selecionado=ano_selecionado
        )
    else:
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))


@user_bp.route('/user/api/dados-empresa/<int:empresa_id>/<int:ano>')
def api_dados_empresa_user(empresa_id, ano):
    """API compatível com template do admin - retorna dados organizados por subgrupo"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        return jsonify({"error": "Não autorizado"}), 403

    from models.user_manager import UserManager
    from models.company_manager import CompanyManager

    # SEGURANÇA: Verificar se o usuário tem acesso a esta empresa
    empresa_id_session = session.get('empresa_id')
    if empresa_id != empresa_id_session:
        return jsonify({"error": "Acesso negado a esta empresa"}), 403

    user_manager = UserManager()
    user_data = user_manager.find_user_by_email(session.get('user_email'))
    user_manager.close()

    if not user_data:
        return jsonify({"error": "Usuário não encontrado"}), 404

    company_manager = CompanyManager()
    data_results = company_manager.buscar_dados_empresa(empresa_id, ano)
    company_manager.close()

    # Organizar dados por SUBGRUPO dentro de cada grupo de viabilidade (MESMA LÓGICA DO ADMIN)
    dados_organizados = {}

    # Processar TbItens
    if data_results and data_results.get('TbItens'):
        for item in data_results['TbItens']:
            grupo = item[0]
            subgrupo = item[1]
            descricao = item[2]
            percentual = float(item[3]) if item[3] else 0
            valor = float(item[4]) if item[4] else 0

            if grupo not in dados_organizados:
                dados_organizados[grupo] = {}

            if subgrupo not in dados_organizados[grupo]:
                dados_organizados[grupo][subgrupo] = []

            dados_organizados[grupo][subgrupo].append({
                "descricao": descricao,
                "percentual": percentual,
                "valor": valor
            })

    # Processar Investimentos (TODOS OS CAMPOS)
    if data_results and data_results.get('TbItensInvestimentos'):
        for item in data_results['TbItensInvestimentos']:
            grupo = item[0]
            subgrupo = item[1]
            descricao = item[2]
            parcela = float(item[3]) if len(item) > 3 and item[3] else 0
            juros = float(item[4]) if len(item) > 4 and item[4] else 0
            total = float(item[5]) if len(item) > 5 and item[5] else 0

            if grupo not in dados_organizados:
                dados_organizados[grupo] = {}

            if subgrupo not in dados_organizados[grupo]:
                dados_organizados[grupo][subgrupo] = []

            dados_organizados[grupo][subgrupo].append({
                "descricao": descricao,
                "parcela": parcela,
                "juros": juros,
                "valor": total,
                "percentual": 0
            })

    # Processar Dívidas (TODOS OS CAMPOS)
    if data_results and data_results.get('TbItensDividas'):
        for item in data_results['TbItensDividas']:
            grupo = item[0]
            subgrupo = item[1]
            descricao = item[2]
            parcela = float(item[3]) if len(item) > 3 and item[3] else 0
            juros = float(item[4]) if len(item) > 4 and item[4] else 0
            total = float(item[5]) if len(item) > 5 and item[5] else 0

            if grupo not in dados_organizados:
                dados_organizados[grupo] = {}

            if subgrupo not in dados_organizados[grupo]:
                dados_organizados[grupo][subgrupo] = []

            dados_organizados[grupo][subgrupo].append({
                "descricao": descricao,
                "parcela": parcela,
                "juros": juros,
                "valor": total,
                "percentual": 0
            })

    # Processar Investimento Geral
    if data_results and data_results.get('TbItensInvestimentoGeral'):
        for item in data_results['TbItensInvestimentoGeral']:
            grupo = item[0]
            subgrupo = item[1]
            descricao = item[2]
            valor = float(item[3]) if len(item) > 3 and item[3] else 0

            if grupo not in dados_organizados:
                dados_organizados[grupo] = {}

            if subgrupo not in dados_organizados[grupo]:
                dados_organizados[grupo][subgrupo] = []

            dados_organizados[grupo][subgrupo].append({
                "descricao": descricao,
                "valor": valor,
                "percentual": 0
            })

    # Processar Gastos Operacionais (COM NOME DIFERENCIADO)
    if data_results and data_results.get('TbItensGastosOperacionais'):
        for item in data_results['TbItensGastosOperacionais']:
            grupo = item[0]
            subgrupo_original = item[1]  # Pode ser "GastosOperacionais"
            descricao = item[2]
            custo_km = float(item[3]) if len(item) > 3 and item[3] else 0
            custo_mensal = float(item[4]) if len(item) > 4 and item[4] else 0

            # Renomear subgrupo para "Gastos Operacionais Veículos" para diferenciar
            subgrupo = 'Gastos Operacionais Veículos'

            if grupo not in dados_organizados:
                dados_organizados[grupo] = {}

            if subgrupo not in dados_organizados[grupo]:
                dados_organizados[grupo][subgrupo] = []

            dados_organizados[grupo][subgrupo].append({
                "descricao": descricao,
                "custo_km": custo_km,
                "valor": custo_mensal,
                "percentual": 0
            })

    return jsonify({
        "ano": ano,
        "dados": dados_organizados
    })


@user_bp.route('/user/api/dados/<int:ano>')
def api_dados_ano(ano):
    """API para retornar dados de um ano específico organizados por grupo de viabilidade"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        return jsonify({"error": "Não autorizado"}), 403

    from models.user_manager import UserManager
    from models.company_manager import CompanyManager

    # Verifica se tem empresa selecionada
    if 'empresa_id' not in session:
        return jsonify({"error": "Empresa não selecionada"}), 400

    user_manager = UserManager()
    user_data = user_manager.find_user_by_email(session.get('user_email'))
    user_manager.close()

    if not user_data:
        return jsonify({"error": "Usuário não encontrado"}), 404

    empresa_id = session.get('empresa_id')

    company_manager = CompanyManager()
    data_results = company_manager.buscar_dados_empresa(
        empresa_id,
        ano
    )
    company_manager.close()

    # Organizar dados por grupo de viabilidade
    dados_por_grupo = {
        "Viabilidade Real": {},
        "Viabilidade PE": {},
        "Viabilidade Ideal": {}
    }

    # Processar TbItens
    if data_results and data_results.get('TbItens'):
        for item in data_results['TbItens']:
            grupo = item[0]  # 'Viabilidade Real', 'Viabilidade PE', ou 'Viabilidade Ideal'
            subgrupo = item[1]  # 'Geral', 'Receita', 'Controle', etc
            descricao = item[2]
            percentual = float(item[3]) if item[3] else 0
            valor = float(item[4]) if item[4] else 0

            if grupo not in dados_por_grupo:
                continue

            # Criar chave única: subgrupo + descrição
            chave = f"{subgrupo} - {descricao}"

            if chave not in dados_por_grupo[grupo]:
                dados_por_grupo[grupo][chave] = 0

            dados_por_grupo[grupo][chave] += valor

    # Processar outras tabelas (Investimentos, Dívidas, etc) se necessário
    for tabela in ['TbItensInvestimentos', 'TbItensDividas', 'TbItensInvestimentoGeral', 'TbItensGastosOperacionais']:
        if data_results and data_results.get(tabela):
            for item in data_results[tabela]:
                grupo = item[0]
                subgrupo = item[1]
                descricao = item[2]

                # Pegar o valor adequado dependendo da tabela
                if tabela in ['TbItensInvestimentos', 'TbItensDividas']:
                    valor = float(item[5]) if len(item) > 5 and item[5] else 0  # valor_total_parc
                elif tabela == 'TbItensInvestimentoGeral':
                    valor = float(item[3]) if len(item) > 3 and item[3] else 0
                elif tabela == 'TbItensGastosOperacionais':
                    valor = float(item[4]) if len(item) > 4 and item[4] else 0  # valor_mensal
                else:
                    continue

                if grupo not in dados_por_grupo:
                    continue

                chave = f"{subgrupo} - {descricao}"

                if chave not in dados_por_grupo[grupo]:
                    dados_por_grupo[grupo][chave] = 0

                dados_por_grupo[grupo][chave] += valor

    return jsonify({
        "ano": ano,
        "grupos": dados_por_grupo
    })


@user_bp.route('/user/bpo')
def visualizar_bpo():
    """Página de visualização do BPO da empresa selecionada"""
    if 'user_email' in session and session.get('user_role') == 'user':
        from models.user_manager import UserManager

        # Verifica se tem empresa selecionada
        if 'empresa_id' not in session:
            return redirect(url_for('user.selecionar_empresa'))

        # Pega informações do usuário logado
        user_manager = UserManager()
        user_data = user_manager.find_user_by_email(session.get('user_email'))
        user_manager.close()

        if not user_data:
            flash("Erro ao carregar dados do usuário.", "danger")
            return redirect(url_for('index.login'))

        empresa_id = session.get('empresa_id')
        empresa_nome = session.get('empresa_nome')

        # Criar objeto empresa compatível com o template do admin
        empresa = {'nome': empresa_nome}

        # Buscar meses disponíveis para o calendário
        from models.company_manager import CompanyManager
        company_manager = CompanyManager()
        meses_disponiveis = company_manager.listar_meses_bpo_empresa(empresa_id)
        company_manager.close()

        # Renderizar dashboard BPO do usuário
        return render_template(
            'user/dashboard_bpo.html',
            user=user_data,
            empresa=empresa,
            empresa_nome=empresa_nome,
            empresa_id=empresa_id,
            meses_disponiveis=meses_disponiveis,
            is_user_view=True  # Flag para o template saber que é visualização de usuário
        )
    else:
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))


@user_bp.route('/user/trocar-empresa')
def trocar_empresa():
    """Permite o usuário trocar de empresa"""
    if 'user_email' in session and session.get('user_role') == 'user':
        # Remove empresa da sessão e redireciona para seleção
        session.pop('empresa_id', None)
        session.pop('empresa_nome', None)
        return redirect(url_for('user.selecionar_empresa'))
    else:
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))


@user_bp.route('/user/api/dados-bpo/<int:empresa_id>')
def api_dados_bpo_user(empresa_id):
    """API compatível com template do admin - retorna dados BPO processados"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        return jsonify({"error": "Não autorizado"}), 403

    # SEGURANÇA: Verificar se o usuário tem acesso a esta empresa
    empresa_id_session = session.get('empresa_id')
    if empresa_id != empresa_id_session:
        return jsonify({"error": "Acesso negado a esta empresa"}), 403

    ano_inicio = int(request.args.get('ano_inicio', 2025))
    mes_inicio = int(request.args.get('mes_inicio', 1))
    ano_fim = int(request.args.get('ano_fim', 2025))
    mes_fim = int(request.args.get('mes_fim', 12))
    tipo_dre = request.args.get('tipo_dre', 'fluxo_caixa')

    from models.company_manager import CompanyManager
    company_manager = CompanyManager()

    # Buscar todos os meses
    meses_data = []
    ano_atual = ano_inicio
    mes_atual = mes_inicio

    while (ano_atual < ano_fim) or (ano_atual == ano_fim and mes_atual <= mes_fim):
        dados = company_manager.buscar_dados_bpo_empresa(empresa_id, ano_atual, mes_atual)
        if dados:
            meses_data.append({
                'ano': ano_atual,
                'mes': mes_atual,
                'dados': dados['dados']
            })
        mes_atual += 1
        if mes_atual > 12:
            mes_atual = 1
            ano_atual += 1

    company_manager.close()

    # Inicializar totais acumulados
    totais = {
        'fluxo_caixa': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real_mp': {'receita': 0, 'despesa': 0, 'geral': 0}
    }

    # Totais de orçamento (para média prevista)
    totais_orcamento = {
        'fluxo_caixa': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real_mp': {'receita': 0, 'despesa': 0, 'geral': 0}
    }

    # Arrays para gráficos (por mês, do DRE selecionado)
    labels_meses = []
    receitas_mensais = []
    despesas_mensais = []
    gerais_mensais = []

    # Nomes dos meses
    nomes_meses = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }

    for mes_data in meses_data:
        mes_num = mes_data['mes']
        ano = mes_data['ano']
        dados = mes_data['dados']

        # Label para gráfico (formato: Janeiro/25)
        nome_mes = nomes_meses.get(mes_num, str(mes_num))
        ano_curto = str(ano)[-2:]
        labels_meses.append(f"{nome_mes}/{ano_curto}")

        # Extrair totais_calculados
        totais_calculados = dados.get('totais_calculados', {})

        if not totais_calculados or totais_calculados == {}:
            receitas_mensais.append(0)
            despesas_mensais.append(0)
            gerais_mensais.append(0)
            continue

        # Variáveis para gráfico deste mês
        receita_grafico = 0
        despesa_grafico = 0
        geral_grafico = 0

        # Processar cada cenário (fluxo_caixa, real, real_mp)
        for cenario_key in ['fluxo_caixa', 'real', 'real_mp']:
            cenario_data = totais_calculados.get(cenario_key, {})

            if not cenario_data or not isinstance(cenario_data, dict):
                continue

            mes_dados = cenario_data.get(mes_num, cenario_data.get(str(mes_num), {}))

            if mes_dados and isinstance(mes_dados, dict):
                # Extrair valores realizados
                realizado = mes_dados.get('realizado', {})
                if isinstance(realizado, dict):
                    receita = realizado.get('receita', 0) or 0
                    despesa = realizado.get('despesa', 0) or 0
                    geral = realizado.get('geral', 0) or 0

                    # RECALCULAR DESPESA DO REAL_MP SE EXISTIR PERCENTUAL MANUAL
                    if cenario_key == 'real_mp':
                        percentual_mp_manual = dados.get('percentual_mp_manual')
                        if percentual_mp_manual is not None:
                            # Pegar a despesa do "Resultado Real"
                            despesa_real = totais_calculados.get('real', {}).get(mes_num, totais_calculados.get('real', {}).get(str(mes_num), {})).get('realizado', {}).get('despesa', 0) or 0

                            # Buscar o valor do item 2.08 (Matéria Prima)
                            valor_materia_prima = 0
                            itens = dados.get('itens_hierarquicos', [])
                            items_to_process = itens if isinstance(itens, list) else itens.items()

                            for item in items_to_process:
                                if isinstance(itens, list):
                                    codigo = item.get('codigo', '')
                                    item_data = item
                                else:
                                    codigo, item_data = item

                                if codigo == '2.08':
                                    dados_mensais = item_data.get('dados_mensais', [])
                                    if dados_mensais and len(dados_mensais) > 0:
                                        mes_atual_dados = dados_mensais[0]
                                        valor_materia_prima = mes_atual_dados.get('valor_realizado', 0) or 0
                                    break

                            # Nova fórmula: Despesa_Real - Matéria_Prima + (Receita × Percentual)
                            despesa_recalculada = despesa_real - valor_materia_prima + ((percentual_mp_manual / 100) * receita)
                            despesa = despesa_recalculada
                            geral = receita - despesa

                    totais[cenario_key]['receita'] += receita
                    totais[cenario_key]['despesa'] += despesa
                    totais[cenario_key]['geral'] += geral

                    if cenario_key == tipo_dre:
                        receita_grafico = receita
                        despesa_grafico = despesa
                        geral_grafico = geral

                # Extrair valores de orçamento
                orcamento = mes_dados.get('orcamento', {})
                if isinstance(orcamento, dict):
                    receita_orc = orcamento.get('receita', 0) or 0
                    despesa_orc = orcamento.get('despesa', 0) or 0
                    geral_orc = orcamento.get('geral', 0) or 0

                    totais_orcamento[cenario_key]['receita'] += receita_orc
                    totais_orcamento[cenario_key]['despesa'] += despesa_orc
                    totais_orcamento[cenario_key]['geral'] += geral_orc

        # Adicionar aos arrays do gráfico
        receitas_mensais.append(receita_grafico)
        despesas_mensais.append(despesa_grafico)
        gerais_mensais.append(geral_grafico)

    # Processar categorias de despesa (itens 2.0X)
    categorias_despesa = {}
    total_receita_orcado = 0

    try:
        for mes_data in meses_data:
            dados = mes_data['dados']
            itens = dados.get('itens_hierarquicos', [])

            items_to_process = itens if isinstance(itens, list) else itens.items()

            for item in items_to_process:
                if isinstance(itens, list):
                    codigo = item.get('codigo', '')
                    item_data = item
                else:
                    codigo, item_data = item

                if codigo.startswith('2.') and len(codigo.split('.')) == 2 and codigo.split('.')[0] == '2' and codigo.split('.')[1].startswith('0'):
                    if codigo not in categorias_despesa:
                        categorias_despesa[codigo] = {
                            'nome': item_data.get('nome', codigo),
                            'orcado': 0,
                            'realizado': 0
                        }

                    dados_mensais = item_data.get('dados_mensais', [])
                    if dados_mensais and len(dados_mensais) > 0:
                        mes_atual_dados = dados_mensais[0]
                        orcado_val = mes_atual_dados.get('valor_orcado', 0) or 0
                        realizado_val = mes_atual_dados.get('valor_realizado', 0) or 0

                        if categorias_despesa[codigo]['orcado'] == 0:
                            categorias_despesa[codigo]['orcado'] = orcado_val

                        categorias_despesa[codigo]['realizado'] += realizado_val

            # Pegar total receita orçado
            if total_receita_orcado == 0:
                totais_calc = dados.get('totais_calculados', {})
                cenario_fc = totais_calc.get(tipo_dre, {})
                if cenario_fc and isinstance(cenario_fc, dict) and len(cenario_fc) > 0:
                    try:
                        primeiro_mes = list(cenario_fc.keys())[0]
                        mes_info = cenario_fc.get(primeiro_mes, {})
                        if isinstance(mes_info, dict):
                            orcamento_info = mes_info.get('orcamento', {})
                            if isinstance(orcamento_info, dict):
                                total_receita_orcado = orcamento_info.get('receita', 0) or 0
                    except (IndexError, KeyError, TypeError):
                        pass

        # Calcular médias e diferenças
        num_meses = len(meses_data)
        for codigo in categorias_despesa:
            cat = categorias_despesa[codigo]
            cat['realizado'] = cat['realizado'] / num_meses if num_meses > 0 else 0
            cat['diferenca'] = cat['orcado'] - cat['realizado']

    except Exception:
        categorias_despesa = {}
        total_receita_orcado = 0

    # Processar categorias de receita (itens 1.0X)
    categorias_receita = {}

    try:
        for mes_data in meses_data:
            dados = mes_data['dados']
            itens = dados.get('itens_hierarquicos', [])

            items_to_process = itens if isinstance(itens, list) else itens.items()

            for item in items_to_process:
                if isinstance(itens, list):
                    codigo = item.get('codigo', '')
                    item_data = item
                else:
                    codigo, item_data = item

                if codigo.startswith('1.') and len(codigo.split('.')) == 2 and codigo.split('.')[0] == '1' and codigo.split('.')[1].startswith('0'):
                    if codigo not in categorias_receita:
                        categorias_receita[codigo] = {
                            'nome': item_data.get('nome', codigo),
                            'orcado': 0,
                            'realizado': 0
                        }

                    dados_mensais = item_data.get('dados_mensais', [])
                    if dados_mensais and len(dados_mensais) > 0:
                        mes_atual_dados = dados_mensais[0]
                        orcado_val = mes_atual_dados.get('valor_orcado', 0) or 0
                        realizado_val = mes_atual_dados.get('valor_realizado', 0) or 0

                        if categorias_receita[codigo]['orcado'] == 0:
                            categorias_receita[codigo]['orcado'] = orcado_val

                        categorias_receita[codigo]['realizado'] += realizado_val

        # Calcular médias e diferenças
        num_meses = len(meses_data)
        for codigo in categorias_receita:
            cat = categorias_receita[codigo]
            cat['realizado'] = cat['realizado'] / num_meses if num_meses > 0 else 0
            cat['diferenca'] = cat['orcado'] - cat['realizado']

    except Exception:
        categorias_receita = {}

    return jsonify({
        'totais_acumulados': totais,
        'totais_orcamento': totais_orcamento,
        'num_meses': len(meses_data),
        'meses': labels_meses,
        'receitas': receitas_mensais,
        'despesas': despesas_mensais,
        'gerais': gerais_mensais,
        'categorias_despesa': categorias_despesa,
        'categorias_receita': categorias_receita,
        'total_receita_orcado': total_receita_orcado
    })


@user_bp.route('/user/consultar-bpo', methods=['GET', 'POST'])
def consultar_dados_bpo():
    """Consulta dados de BPO em formato de tabela para a empresa do usuário"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        flash("Acesso negado. Faça login como usuário.", "danger")
        return redirect(url_for('index.login'))

    # Verifica se tem empresa selecionada
    if 'empresa_id' not in session:
        return redirect(url_for('user.selecionar_empresa'))

    from models.user_manager import UserManager
    from models.company_manager import CompanyManager

    # Pega informações do usuário logado
    user_manager = UserManager()
    user_data = user_manager.find_user_by_email(session.get('user_email'))
    user_manager.close()

    if not user_data:
        flash("Erro ao carregar dados do usuário.", "danger")
        return redirect(url_for('index.login'))

    empresa_id = session.get('empresa_id')
    empresa_nome = session.get('empresa_nome')

    data_results = None
    ano_selecionado = None
    mes_selecionado = None

    if request.method == 'POST':
        ano_selecionado = request.form.get('ano')
        mes_selecionado = request.form.get('mes')

        if ano_selecionado and mes_selecionado:
            company_manager = CompanyManager()
            data_results = company_manager.buscar_dados_bpo_empresa(
                empresa_id,
                int(ano_selecionado),
                int(mes_selecionado)
            )
            company_manager.close()

    return render_template(
        'user/consultar_bpo.html',
        user=user_data,
        empresa_id=empresa_id,
        empresa_nome=empresa_nome,
        data_results=data_results,
        ano_selecionado=ano_selecionado,
        mes_selecionado=mes_selecionado
    )


@user_bp.route('/user/gerar_relatorio_bpo')
def gerar_relatorio_bpo():
    """Gera relatório Excel do dashboard BPO para usuário"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        return "Acesso negado", 403

    # Pega empresa_id da sessão do usuário
    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return "Empresa não selecionada", 400

    from models.company_manager import CompanyManager
    from datetime import datetime
    from flask import make_response
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    # Buscar dados da empresa
    company_manager = CompanyManager()
    empresa = company_manager.buscar_empresa_por_id(empresa_id)

    if not empresa:
        company_manager.close()
        return "Empresa não encontrada", 404

    # Pegar parâmetros do filtro
    ano_inicio = int(request.args.get('ano_inicio', 2025))
    mes_inicio = int(request.args.get('mes_inicio', 1))
    ano_fim = int(request.args.get('ano_fim', 2025))
    mes_fim = int(request.args.get('mes_fim', 12))
    tipo_dre = request.args.get('tipo_dre', 'fluxo_caixa')

    # ========== PROCESSAR DADOS BPO (mesma lógica da API) ==========

    # Buscar todos os meses
    meses_data = []
    ano_atual = ano_inicio
    mes_atual = mes_inicio

    while (ano_atual < ano_fim) or (ano_atual == ano_fim and mes_atual <= mes_fim):
        dados = company_manager.buscar_dados_bpo_empresa(empresa_id, ano_atual, mes_atual)
        if dados:
            meses_data.append({
                'ano': ano_atual,
                'mes': mes_atual,
                'dados': dados['dados']
            })
        mes_atual += 1
        if mes_atual > 12:
            mes_atual = 1
            ano_atual += 1

    # Inicializar totais acumulados
    totais = {
        'fluxo_caixa': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real_mp': {'receita': 0, 'despesa': 0, 'geral': 0}
    }

    # Totais de orçamento (para média prevista)
    totais_orcamento = {
        'fluxo_caixa': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real': {'receita': 0, 'despesa': 0, 'geral': 0},
        'real_mp': {'receita': 0, 'despesa': 0, 'geral': 0}
    }

    # Arrays para gráficos (por mês, do DRE selecionado)
    labels_meses = []
    receitas_mensais = []
    despesas_mensais = []
    gerais_mensais = []

    # Nomes dos meses
    nomes_meses = {
        1: 'Janeiro', 2: 'Fevereiro', 3: 'Março', 4: 'Abril',
        5: 'Maio', 6: 'Junho', 7: 'Julho', 8: 'Agosto',
        9: 'Setembro', 10: 'Outubro', 11: 'Novembro', 12: 'Dezembro'
    }

    for mes_data in meses_data:
        mes_num = mes_data['mes']
        ano = mes_data['ano']
        dados = mes_data['dados']

        # Label para gráfico (formato: Janeiro/25)
        nome_mes = nomes_meses.get(mes_num, str(mes_num))
        ano_curto = str(ano)[-2:]
        labels_meses.append(f"{nome_mes}/{ano_curto}")

        # Extrair totais_calculados
        totais_calculados = dados.get('totais_calculados', {})

        if not totais_calculados or totais_calculados == {}:
            receitas_mensais.append(0)
            despesas_mensais.append(0)
            gerais_mensais.append(0)
            continue

        # Variáveis para gráfico deste mês
        receita_grafico = 0
        despesa_grafico = 0
        geral_grafico = 0

        # Processar cada cenário (fluxo_caixa, real, real_mp)
        for cenario_key in ['fluxo_caixa', 'real', 'real_mp']:
            cenario_data = totais_calculados.get(cenario_key, {})

            if not cenario_data or not isinstance(cenario_data, dict):
                continue

            # Pegar dados do mês
            mes_dados = cenario_data.get(mes_num, cenario_data.get(str(mes_num), {}))

            if mes_dados and isinstance(mes_dados, dict):
                # Extrair valores realizados
                realizado = mes_dados.get('realizado', {})
                if isinstance(realizado, dict):
                    receita = realizado.get('receita', 0) or 0
                    despesa = realizado.get('despesa', 0) or 0
                    geral = realizado.get('geral', 0) or 0

                    # RECALCULAR DESPESA DO REAL_MP SE EXISTIR PERCENTUAL MANUAL
                    if cenario_key == 'real_mp':
                        percentual_mp_manual = dados.get('percentual_mp_manual')
                        if percentual_mp_manual is not None:
                            # Pegar a despesa do "Resultado Real"
                            despesa_real = totais_calculados.get('real', {}).get(mes_num, totais_calculados.get('real', {}).get(str(mes_num), {})).get('realizado', {}).get('despesa', 0) or 0

                            # Buscar o valor do item 2.08 (Matéria Prima)
                            valor_materia_prima = 0
                            itens = dados.get('itens_hierarquicos', [])

                            # Itens pode ser lista ou dict
                            items_to_process = itens if isinstance(itens, list) else itens.items()

                            for item in items_to_process:
                                if isinstance(itens, list):
                                    codigo = item.get('codigo', '')
                                    item_data = item
                                else:
                                    codigo, item_data = item

                                # Verificar se é o item 2.08
                                if codigo == '2.08':
                                    dados_mensais = item_data.get('dados_mensais', [])
                                    if dados_mensais and len(dados_mensais) > 0:
                                        mes_atual_dados = dados_mensais[0]
                                        valor_materia_prima = mes_atual_dados.get('valor_realizado', 0) or 0
                                    break

                            # Nova fórmula: Despesa_Real - Matéria_Prima + (Receita × Percentual)
                            despesa_recalculada = despesa_real - valor_materia_prima + ((percentual_mp_manual / 100) * receita)
                            despesa = despesa_recalculada
                            geral = receita - despesa

                    # Acumular totais
                    totais[cenario_key]['receita'] += receita
                    totais[cenario_key]['despesa'] += despesa
                    totais[cenario_key]['geral'] += geral

                    # Se é o DRE selecionado, guardar para gráfico
                    if cenario_key == tipo_dre:
                        receita_grafico = receita
                        despesa_grafico = despesa
                        geral_grafico = geral

                # Extrair valores de orçamento
                orcamento = mes_dados.get('orcamento', {})
                if isinstance(orcamento, dict):
                    receita_orc = orcamento.get('receita', 0) or 0
                    despesa_orc = orcamento.get('despesa', 0) or 0
                    geral_orc = orcamento.get('geral', 0) or 0

                    # Acumular orçamento
                    totais_orcamento[cenario_key]['receita'] += receita_orc
                    totais_orcamento[cenario_key]['despesa'] += despesa_orc
                    totais_orcamento[cenario_key]['geral'] += geral_orc

        # Adicionar aos arrays do gráfico
        receitas_mensais.append(receita_grafico)
        despesas_mensais.append(despesa_grafico)
        gerais_mensais.append(geral_grafico)

    # Processar categorias de despesa (itens 2.0X)
    categorias_despesa = {}

    for mes_data in meses_data:
        dados = mes_data['dados']
        itens = dados.get('itens_hierarquicos', [])

        items_to_process = itens if isinstance(itens, list) else itens.items()

        for item in items_to_process:
            if isinstance(itens, list):
                codigo = item.get('codigo', '')
                item_data = item
            else:
                codigo, item_data = item

            # Filtrar apenas itens 2.0X
            if codigo.startswith('2.') and len(codigo.split('.')) == 2 and codigo.split('.')[0] == '2' and codigo.split('.')[1].startswith('0'):
                if codigo not in categorias_despesa:
                    categorias_despesa[codigo] = {
                        'nome': item_data.get('nome', codigo),
                        'orcado': 0,
                        'realizado': 0
                    }

                dados_mensais = item_data.get('dados_mensais', [])
                if dados_mensais and len(dados_mensais) > 0:
                    mes_atual_dados = dados_mensais[0]
                    orcado_val = mes_atual_dados.get('valor_orcado', 0) or 0
                    realizado_val = mes_atual_dados.get('valor_realizado', 0) or 0

                    if categorias_despesa[codigo]['orcado'] == 0:
                        categorias_despesa[codigo]['orcado'] = orcado_val

                    categorias_despesa[codigo]['realizado'] += realizado_val

    # Calcular médias de categorias de despesa
    num_meses = len(meses_data)
    for codigo in categorias_despesa:
        cat = categorias_despesa[codigo]
        cat['realizado'] = cat['realizado'] / num_meses if num_meses > 0 else 0
        cat['diferenca'] = cat['realizado'] - cat['orcado']

    # Processar categorias de receita (itens 1.0X)
    categorias_receita = {}

    for mes_data in meses_data:
        dados = mes_data['dados']
        itens = dados.get('itens_hierarquicos', [])

        items_to_process = itens if isinstance(itens, list) else itens.items()

        for item in items_to_process:
            if isinstance(itens, list):
                codigo = item.get('codigo', '')
                item_data = item
            else:
                codigo, item_data = item

            # Filtrar apenas itens 1.0X
            if codigo.startswith('1.') and len(codigo.split('.')) == 2 and codigo.split('.')[0] == '1' and codigo.split('.')[1].startswith('0'):
                if codigo not in categorias_receita:
                    categorias_receita[codigo] = {
                        'nome': item_data.get('nome', codigo),
                        'orcado': 0,
                        'realizado': 0
                    }

                dados_mensais = item_data.get('dados_mensais', [])
                if dados_mensais and len(dados_mensais) > 0:
                    mes_atual_dados = dados_mensais[0]
                    orcado_val = mes_atual_dados.get('valor_orcado', 0) or 0
                    realizado_val = mes_atual_dados.get('valor_realizado', 0) or 0

                    if categorias_receita[codigo]['orcado'] == 0:
                        categorias_receita[codigo]['orcado'] = orcado_val

                    categorias_receita[codigo]['realizado'] += realizado_val

    # Calcular médias de categorias de receita
    for codigo in categorias_receita:
        cat = categorias_receita[codigo]
        cat['realizado'] = cat['realizado'] / num_meses if num_meses > 0 else 0
        cat['diferenca'] = cat['realizado'] - cat['orcado']

    company_manager.close()

    # ========== CRIAR EXCEL COM GRÁFICOS ==========
    from openpyxl.chart import LineChart, PieChart

    wb = Workbook()
    ws = wb.active
    ws.title = "Gráficos BPO"

    # Estilos
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=16)

    # Cabeçalho
    ws['A1'] = 'Relatório BPO - Análise Gráfica'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f"Empresa: {empresa['nome']}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A3'] = f"Período: {nomes_meses[mes_inicio]}/{ano_inicio} - {nomes_meses[mes_fim]}/{ano_fim}"
    ws['A3'].font = Font(size=10)
    ws['A4'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A4'].font = Font(size=9, italic=True)

    # Dados ocultos para os gráficos (coluna A-D, começando linha 6)
    row = 6

    # ==== GRÁFICO 1: COMPARATIVO DOS 3 DREs ====
    ws.cell(row, 1, "DRE")
    ws.cell(row, 2, "Resultado")
    row += 1
    for dre_nome, dre_key in [('Fluxo Caixa', 'fluxo_caixa'), ('Real', 'real'), ('Real+MP', 'real_mp')]:
        ws.cell(row, 1, dre_nome)
        ws.cell(row, 2, totais[dre_key]['geral'])
        row += 1

    chart1 = BarChart()
    chart1.title = "Comparativo de Resultado - 3 DREs"
    chart1.y_axis.title = 'Resultado (R$)'
    chart1.style = 11
    data = Reference(ws, min_col=2, min_row=7, max_row=9)
    cats = Reference(ws, min_col=1, min_row=7, max_row=9)
    chart1.add_data(data, titles_from_data=False)
    chart1.set_categories(cats)
    chart1.height = 10
    chart1.width = 18
    ws.add_chart(chart1, 'J7')

    # ==== GRÁFICO 2: EVOLUÇÃO MENSAL (Receita x Despesa x Resultado) ====
    row = 6
    col_offset = 5  # Coluna E
    ws.cell(row, col_offset, "Mês")
    ws.cell(row, col_offset+1, "Receita")
    ws.cell(row, col_offset+2, "Despesa")
    ws.cell(row, col_offset+3, "Resultado")
    row += 1

    for i, label in enumerate(labels_meses):
        ws.cell(row, col_offset, label)
        ws.cell(row, col_offset+1, receitas_mensais[i])
        ws.cell(row, col_offset+2, despesas_mensais[i])
        ws.cell(row, col_offset+3, gerais_mensais[i])
        row += 1

    chart2 = LineChart()
    chart2.title = f"Evolução Mensal - {tipo_dre.replace('_', ' ').title()}"
    chart2.y_axis.title = 'Valor (R$)'
    chart2.x_axis.title = 'Mês'
    chart2.style = 12
    max_row_chart2 = 6 + len(labels_meses)
    data = Reference(ws, min_col=col_offset+1, min_row=6, max_row=max_row_chart2, max_col=col_offset+3)
    cats = Reference(ws, min_col=col_offset, min_row=7, max_row=max_row_chart2)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 12
    chart2.width = 22
    ws.add_chart(chart2, 'A12')

    # ==== GRÁFICO 3: PIZZA CATEGORIAS DE RECEITA ====
    row_start_receita = 6 + len(labels_meses) + 3
    row = row_start_receita
    ws.cell(row, 1, "Categoria Receita")
    ws.cell(row, 2, "Valor")
    row += 1

    for codigo in sorted(categorias_receita.keys()):
        cat = categorias_receita[codigo]
        ws.cell(row, 1, cat['nome'])
        ws.cell(row, 2, cat['realizado'])
        row += 1

    if len(categorias_receita) > 0:
        chart3 = PieChart()
        chart3.title = "Distribuição de Receitas por Categoria"
        chart3.style = 10
        data = Reference(ws, min_col=2, min_row=row_start_receita+1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=row_start_receita+1, max_row=row-1)
        chart3.add_data(data, titles_from_data=False)
        chart3.set_categories(cats)
        chart3.height = 12
        chart3.width = 16
        ws.add_chart(chart3, 'J22')

    # ==== GRÁFICO 4: PIZZA CATEGORIAS DE DESPESA ====
    row_start_despesa = row + 2
    row = row_start_despesa
    ws.cell(row, 1, "Categoria Despesa")
    ws.cell(row, 2, "Valor")
    row += 1

    for codigo in sorted(categorias_despesa.keys()):
        cat = categorias_despesa[codigo]
        ws.cell(row, 1, cat['nome'])
        ws.cell(row, 2, cat['realizado'])
        row += 1

    if len(categorias_despesa) > 0:
        chart4 = PieChart()
        chart4.title = "Distribuição de Despesas por Categoria"
        chart4.style = 10
        data = Reference(ws, min_col=2, min_row=row_start_despesa+1, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=row_start_despesa+1, max_row=row-1)
        chart4.add_data(data, titles_from_data=False)
        chart4.set_categories(cats)
        chart4.height = 12
        chart4.width = 16
        ws.add_chart(chart4, 'J37')

    # ==== GRÁFICO 5: BARRAS HORIZONTAIS - RECEITA vs DESPESA POR MÊS ====
    row_start_bar = row + 2
    row = row_start_bar
    ws.cell(row, 5, "Mês")
    ws.cell(row, 6, "Receita")
    ws.cell(row, 7, "Despesa")
    row += 1

    for i, label in enumerate(labels_meses):
        ws.cell(row, 5, label)
        ws.cell(row, 6, receitas_mensais[i])
        ws.cell(row, 7, despesas_mensais[i])
        row += 1

    from openpyxl.chart import BarChart as HorizontalBarChart
    chart5 = HorizontalBarChart()
    chart5.type = "bar"
    chart5.title = "Receita vs Despesa Mensal"
    chart5.y_axis.title = 'Mês'
    chart5.x_axis.title = 'Valor (R$)'
    chart5.style = 13
    data = Reference(ws, min_col=6, min_row=row_start_bar, max_row=row-1, max_col=7)
    cats = Reference(ws, min_col=5, min_row=row_start_bar+1, max_row=row-1)
    chart5.add_data(data, titles_from_data=True)
    chart5.set_categories(cats)
    chart5.height = 14
    chart5.width = 20
    ws.add_chart(chart5, 'J52')

    # Ocultar dados (deixar apenas gráficos visíveis)
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].hidden = True

    # Salvar Excel em buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Retornar Excel
    response = make_response(excel_buffer.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Relatorio_BPO_{empresa["nome"]}_{datetime.now().strftime("%Y%m%d")}.xlsx'

    return response

@user_bp.route('/user/gerar_relatorio_viabilidade')
def gerar_relatorio_viabilidade():
    """Gera relatório Excel comparando os 3 grupos de viabilidade para usuário"""
    if not ('user_email' in session and session.get('user_role') == 'user'):
        return "Acesso negado", 403

    empresa_id = session.get('empresa_id')
    if not empresa_id:
        return "Empresa não selecionada", 400

    from models.company_manager import CompanyManager
    from datetime import datetime
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, Reference

    # Buscar dados da empresa
    company_manager = CompanyManager()
    empresa = company_manager.buscar_empresa_por_id(empresa_id)

    if not empresa:
        company_manager.close()
        return "Empresa não encontrada", 404

    # Obter ano selecionado
    ano_selecionado = int(request.args.get('ano_selecionado', datetime.now().year))

    # Buscar dados dos 3 grupos de viabilidade
    dados_completos = company_manager.buscar_dados_empresa(empresa_id, ano_selecionado)

    if not dados_completos:
        company_manager.close()
        return "Sem dados de viabilidade disponíveis para este ano", 404

    # Processar dados dos 3 grupos
    grupos_info = {
        'Viabilidade Real': {'receita': 0, 'despesa': 0, 'resultado': 0, 'subgrupos': {}},
        'Viabilidade PE': {'receita': 0, 'despesa': 0, 'resultado': 0, 'subgrupos': {}},
        'Viabilidade Ideal': {'receita': 0, 'despesa': 0, 'resultado': 0, 'subgrupos': {}}
    }

    # Processar cada grupo
    for grupo_nome, grupo_data in dados_completos.get('dados', {}).items():
        if grupo_nome not in grupos_info:
            continue

        for subgrupo_nome, itens in grupo_data.items():
            total_subgrupo = sum(item.get('valor', 0) for item in itens)

            # Classificar como receita ou despesa baseado no nome do subgrupo
            if subgrupo_nome in ['Receita', 'Geral']:
                grupos_info[grupo_nome]['receita'] += total_subgrupo
            else:
                grupos_info[grupo_nome]['despesa'] += total_subgrupo

            # Armazenar detalhes do subgrupo
            if subgrupo_nome not in grupos_info[grupo_nome]['subgrupos']:
                grupos_info[grupo_nome]['subgrupos'][subgrupo_nome] = []
            grupos_info[grupo_nome]['subgrupos'][subgrupo_nome].extend(itens)

    # Calcular resultados
    for grupo in grupos_info.values():
        grupo['resultado'] = grupo['receita'] - grupo['despesa']

    company_manager.close()

    # ========== CRIAR EXCEL COM GRÁFICOS ==========
    from openpyxl.chart import PieChart

    wb = Workbook()
    ws = wb.active
    ws.title = "Gráficos Viabilidade"

    # Estilos
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=16)

    # Cabeçalho
    ws['A1'] = 'Relatório de Viabilidade - Análise Gráfica'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:H1')
    ws.row_dimensions[1].height = 30

    ws['A2'] = f"Empresa: {empresa['nome']}"
    ws['A2'].font = Font(bold=True, size=11)
    ws['A3'] = f"Ano: {ano_selecionado}"
    ws['A3'].font = Font(size=10)
    ws['A4'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A4'].font = Font(size=9, italic=True)

    # Dados ocultos para os gráficos
    row = 6

    # ==== GRÁFICO 1: COMPARATIVO DE RESULTADOS DOS 3 GRUPOS ====
    ws.cell(row, 1, "Grupo")
    ws.cell(row, 2, "Resultado")
    row += 1
    for grupo_nome in ['Viabilidade Real', 'Viabilidade PE', 'Viabilidade Ideal']:
        grupo = grupos_info[grupo_nome]
        ws.cell(row, 1, grupo_nome.replace('Viabilidade ', ''))
        ws.cell(row, 2, grupo['resultado'])
        row += 1

    chart1 = BarChart()
    chart1.title = "Comparativo de Resultados"
    chart1.y_axis.title = 'Resultado (R$)'
    chart1.style = 11
    data = Reference(ws, min_col=2, min_row=6, max_row=9)
    cats = Reference(ws, min_col=1, min_row=7, max_row=9)
    chart1.add_data(data, titles_from_data=True)
    chart1.set_categories(cats)
    chart1.height = 12
    chart1.width = 18
    ws.add_chart(chart1, 'A7')

    # ==== GRÁFICO 2: RECEITA vs DESPESA POR GRUPO ====
    row = 6
    col_offset = 4  # Coluna D
    ws.cell(row, col_offset, "Grupo")
    ws.cell(row, col_offset+1, "Receita")
    ws.cell(row, col_offset+2, "Despesa")
    row += 1

    for grupo_nome in ['Viabilidade Real', 'Viabilidade PE', 'Viabilidade Ideal']:
        grupo = grupos_info[grupo_nome]
        ws.cell(row, col_offset, grupo_nome.replace('Viabilidade ', ''))
        ws.cell(row, col_offset+1, grupo['receita'])
        ws.cell(row, col_offset+2, grupo['despesa'])
        row += 1

    chart2 = BarChart()
    chart2.title = "Receita vs Despesa por Grupo"
    chart2.y_axis.title = 'Valor (R$)'
    chart2.style = 12
    data = Reference(ws, min_col=col_offset+1, min_row=6, max_row=9, max_col=col_offset+2)
    cats = Reference(ws, min_col=col_offset, min_row=7, max_row=9)
    chart2.add_data(data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.height = 12
    chart2.width = 20
    ws.add_chart(chart2, 'J7')

    # ==== GRÁFICO 3: PIZZA - DISTRIBUIÇÃO DE RECEITAS ====
    row_start_pizza = 12
    row = row_start_pizza
    ws.cell(row, 1, "Grupo")
    ws.cell(row, 2, "Receita")
    row += 1

    for grupo_nome in ['Viabilidade Real', 'Viabilidade PE', 'Viabilidade Ideal']:
        grupo = grupos_info[grupo_nome]
        ws.cell(row, 1, grupo_nome.replace('Viabilidade ', ''))
        ws.cell(row, 2, grupo['receita'])
        row += 1

    chart3 = PieChart()
    chart3.title = "Distribuição de Receitas"
    chart3.style = 10
    data = Reference(ws, min_col=2, min_row=row_start_pizza, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=row_start_pizza+1, max_row=row-1)
    chart3.add_data(data, titles_from_data=True)
    chart3.set_categories(cats)
    chart3.height = 12
    chart3.width = 16
    ws.add_chart(chart3, 'A24')

    # ==== GRÁFICO 4: PIZZA - DISTRIBUIÇÃO DE DESPESAS ====
    row_start_desp = row + 2
    row = row_start_desp
    ws.cell(row, 1, "Grupo")
    ws.cell(row, 2, "Despesa")
    row += 1

    for grupo_nome in ['Viabilidade Real', 'Viabilidade PE', 'Viabilidade Ideal']:
        grupo = grupos_info[grupo_nome]
        ws.cell(row, 1, grupo_nome.replace('Viabilidade ', ''))
        ws.cell(row, 2, grupo['despesa'])
        row += 1

    chart4 = PieChart()
    chart4.title = "Distribuição de Despesas"
    chart4.style = 10
    data = Reference(ws, min_col=2, min_row=row_start_desp, max_row=row-1)
    cats = Reference(ws, min_col=1, min_row=row_start_desp+1, max_row=row-1)
    chart4.add_data(data, titles_from_data=True)
    chart4.set_categories(cats)
    chart4.height = 12
    chart4.width = 16
    ws.add_chart(chart4, 'J24')

    # Ocultar dados (deixar apenas gráficos visíveis)
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].hidden = True

    # Salvar Excel em buffer
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    # Retornar Excel
    response = make_response(excel_buffer.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename=Relatorio_Viabilidade_{empresa["nome"]}_{ano_selecionado}.xlsx'

    return response


@user_bp.route('/user/logout')
def logout():
    """Logout do usuário"""
    session.pop('user_email', None)
    session.pop('user_role', None)
    session.pop('empresa_id', None)
    session.pop('empresa_nome', None)
    flash("Logout realizado com sucesso.", "success")
    return redirect(url_for('index.login'))
