#!/usr/bin/env python3
"""
Script de DEBUG para visualizar estrutura da planilha BPO

Este script mostra EXATAMENTE o que está em cada célula das primeiras linhas,
para podermos identificar a estrutura correta.
"""

import sys
import os
from openpyxl import load_workbook

def debug_planilha(arquivo_path):
    """Exibe estrutura detalhada da planilha"""

    print("\n" + "="*120)
    print(" "*40 + "DEBUG DA PLANILHA BPO")
    print("="*120)

    wb = load_workbook(arquivo_path)
    sheet = wb.active

    print(f"\n📄 Arquivo: {arquivo_path}")
    print(f"📊 Total de linhas: {sheet.max_row}")
    print(f"📊 Total de colunas: {sheet.max_column}")

    # ========================================================================
    # MOSTRAR AS PRIMEIRAS 3 LINHAS COMPLETAS (CABEÇALHOS)
    # ========================================================================
    print("\n" + "="*120)
    print("🔍 PRIMEIRAS 3 LINHAS (CABEÇALHOS)")
    print("="*120)

    for row_num in range(1, 4):
        print(f"\n--- LINHA {row_num} ---")
        for col_num in range(1, min(sheet.max_column + 1, 30)):  # Limitar a 30 colunas
            cell = sheet.cell(row=row_num, column=col_num)
            col_letter = cell.column_letter
            valor = cell.value

            if valor is not None and str(valor).strip():
                print(f"   [{col_letter}{row_num}] = {repr(valor)}")

    # ========================================================================
    # MOSTRAR LINHAS 4 A 10 (DADOS)
    # ========================================================================
    print("\n" + "="*120)
    print("🔍 LINHAS 4 a 10 (DADOS) - TODAS AS COLUNAS")
    print("="*120)

    for row_num in range(4, 11):
        print(f"\n{'='*120}")
        print(f"LINHA {row_num}")
        print(f"{'='*120}")

        # Coluna A
        col_a = sheet.cell(row=row_num, column=1).value
        print(f"  [A{row_num}] Código/Nome: {repr(col_a)}")

        # Mostrar TODAS as colunas com valores
        print(f"\n  Todas as colunas:")
        for col_num in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            col_letter = cell.column_letter
            valor = cell.value

            # Mostrar tipo e valor
            tipo = type(valor).__name__
            print(f"    [{col_letter}{row_num}] ({tipo:8s}) = {repr(valor)}")

    # ========================================================================
    # ANÁLISE DA ESTRUTURA
    # ========================================================================
    print("\n" + "="*120)
    print("📊 ANÁLISE DA ESTRUTURA")
    print("="*120)

    # Contar colunas com dados na linha 4
    linha4_valores = []
    for col_num in range(1, sheet.max_column + 1):
        valor = sheet.cell(row=4, column=col_num).value
        if valor is not None:
            linha4_valores.append((col_num, cell.column_letter, valor))

    print(f"\n✅ Colunas com dados na linha 4: {len(linha4_valores)}")
    print("\nResumo das colunas:")
    for col_num, col_letter, valor in linha4_valores:
        tipo = type(valor).__name__
        print(f"   Coluna {col_num:2d} ({col_letter}) ({tipo:8s}): {repr(valor)[:50]}")

    # Tentar identificar estrutura
    print("\n" + "="*120)
    print("💡 INTERPRETAÇÃO SUGERIDA")
    print("="*120)

    total_colunas = sheet.max_column
    print(f"\nTotal de colunas: {total_colunas}")

    # A estrutura deveria ser:
    # Coluna A (1): Nome/Código
    # Coluna B (2): % Viabilidade
    # Coluna C (3): Valor Viabilidade
    # Colunas D em diante (4+): Dados mensais (4 colunas por mês) + 7 resultados

    colunas_depois_viab = total_colunas - 3
    colunas_antes_resultados = colunas_depois_viab - 7
    num_meses = colunas_antes_resultados // 4

    print(f"\nEstrutura calculada:")
    print(f"  • Coluna 1 (A): Nome/Código")
    print(f"  • Coluna 2 (B): % Viabilidade")
    print(f"  • Coluna 3 (C): Valor Viabilidade")
    print(f"  • Colunas 4-{3 + colunas_antes_resultados} (D-...): {num_meses} meses ({colunas_antes_resultados} colunas)")
    print(f"  • Últimas 7 colunas ({total_colunas - 6}-{total_colunas}): Resultados")

    print(f"\nMapeamento de meses (4 colunas cada):")
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']

    for i in range(num_meses):
        col_inicio = 4 + (i * 4)  # Índice 1-based
        col_fim = col_inicio + 3
        letra_inicio = sheet.cell(row=1, column=col_inicio).column_letter
        letra_fim = sheet.cell(row=1, column=col_fim).column_letter
        mes_nome = meses[i] if i < len(meses) else f'Mês {i+1}'
        print(f"  • {mes_nome}: Colunas {col_inicio}-{col_fim} ({letra_inicio}-{letra_fim})")

    print(f"\nColunas de resultados:")
    nomes_resultados = [
        'Previsão Total',
        'Total Realizado',
        'Diferença Total',
        'Média % Realizado',
        'Média Valor Realizado',
        'Média % Diferença',
        'Média Valor Diferença'
    ]
    col_inicio_resultados = 4 + (num_meses * 4)
    for i, nome in enumerate(nomes_resultados):
        col_num = col_inicio_resultados + i
        letra = sheet.cell(row=1, column=col_num).column_letter
        print(f"  • {nome}: Coluna {col_num} ({letra})")

    print("\n" + "="*120)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("❌ Erro: Forneça o caminho da planilha")
        print(f"\nUso: python3 {sys.argv[0]} planilha.xlsx")
        sys.exit(1)

    arquivo = sys.argv[1]
    if not os.path.exists(arquivo):
        print(f"❌ Arquivo não encontrado: {arquivo}")
        sys.exit(1)

    debug_planilha(arquivo)
    print("\n✅ Debug concluído!\n")
